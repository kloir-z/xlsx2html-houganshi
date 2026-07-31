#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx2html.py — Excel 方眼紙ドキュメントを「見た目そのまま」の単一 HTML に変換する。

方針:
  * 列幅 / 行高 をピクセルに換算し、table-layout:fixed の表として厳密に再現する
    （= 方眼のインデントがそのまま残る）
  * 罫線・塗り・フォント・文字揃え・結合セル・縦書き・回転を CSS で再現
  * 数式は評価済みキャッシュ値を使用（data_only=True）。表示は表示形式に従って整形
  * セル外にはみ出す文字（Excel のオーバーフロー）も再現するため、
    文字は絶対配置の <span> として置き、セル幅に縛られないようにしている
  * 画像・オートシェイプは xl/drawings/*.xml を直接解析し、
    絶対座標のオーバーレイとして重ねる（画像は base64 で埋め込み → 単一ファイル完結）
  * 文字はすべて実テキストなのでコピー可能

使い方:
  python xlsx2html.py input.xlsx [-o output.html] [--sheet 名前] [--gridlines auto|on|off]
"""

from __future__ import annotations

import argparse
import base64
import colorsys
import html
import math
import os
import posixpath
import re
import sys
import warnings
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, datetime, time, timedelta

# openpyxl が読み飛ばす拡張機能（条件付き書式の拡張・ヘッダーフッター等）の
# 警告は変換結果に影響しないので黙らせる
warnings.filterwarnings("ignore", category=UserWarning, module=r"openpyxl.*")

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string, get_column_letter
except ImportError:  # pragma: no cover
    sys.exit("openpyxl が必要です:  pip install openpyxl")


# 日本語版 Excel が使う組み込み表示形式（numFmtId 27-36, 50-58）を openpyxl は
# 知らず、General として扱ってしまう。すると日付セルが datetime に変換されず、
# 46255 のようなシリアル値がそのまま表示される。読み込む前に補っておく。
# これらの ID の意味はロケール依存で、下表は日本語ロケール(LCID 0x411)のもの。
JP_BUILTIN_FORMATS = {
    27: '[$-411]ge.m.d',
    28: '[$-411]ggge"年"m"月"d"日"',
    29: '[$-411]ggge"年"m"月"d"日"',
    30: "m/d/yy",
    31: 'yyyy"年"m"月"d"日"',
    32: 'h"時"mm"分"',
    33: 'h"時"mm"分"ss"秒"',
    34: 'yyyy"年"m"月"',
    35: 'm"月"d"日"',
    36: '[$-411]ge.m.d',
    50: '[$-411]ge.m.d',
    51: '[$-411]ggge"年"m"月"d"日"',
    52: 'yyyy"年"m"月"',
    53: 'm"月"d"日"',
    54: '[$-411]ggge"年"m"月"d"日"',
    55: 'yyyy"年"m"月"',
    56: 'm"月"d"日"',
    57: '[$-411]ge.m.d',
    58: '[$-411]ggge"年"m"月"d"日"',
}

try:
    from openpyxl.styles import numbers as _opx_numbers
    from openpyxl.styles.numbers import is_date_format, is_timedelta_format
    from openpyxl.utils.datetime import WINDOWS_EPOCH, from_excel

    # 辞書を差し替えるのではなく、その場に足す。openpyxl 側は
    # builtin_format_code() 経由でこの辞書を都度引くので、これで効く。
    for _id, _code in JP_BUILTIN_FORMATS.items():
        _opx_numbers.BUILTIN_FORMATS.setdefault(_id, _code)
except Exception:  # pragma: no cover - openpyxl の内部構成が変わった場合
    is_date_format = is_timedelta_format = None
    WINDOWS_EPOCH = datetime(1899, 12, 30)
    from_excel = None


# ---------------------------------------------------------------------------
# 単位換算
# ---------------------------------------------------------------------------

EMU_PER_PX = 9525.0          # 1px = 9525 EMU (96dpi)
MDW = 7                      # 標準フォントの数字1文字幅(px)。Calibri 11 / MS Pゴシック 11 とも 7
DEFAULT_COL_WIDTH = 8.43     # Excel 既定の列幅(文字数)
DEFAULT_ROW_HEIGHT_PT = 18.75
GRID_COLOR = "#e8e8e8"       # Excel の目盛線相当（画面表示のみ、既定では印刷されない）
CELL_PAD = 2                 # セルの左右パディング(px)。Excel 実測相当
INDENT_PX = 8                # インデント1段あたりの px


def col_width_to_px(width: float | None) -> int:
    """Excel の列幅(文字数)→ px。既定 8.43 が 64px になる換算。"""
    if width is None:
        width = DEFAULT_COL_WIDTH
    if width <= 0:
        return 0
    return int(round(width * MDW)) + 5


def pt_to_px(pt: float) -> float:
    return pt * 4.0 / 3.0


def emu_to_px(emu) -> float:
    try:
        return float(emu) / EMU_PER_PX
    except (TypeError, ValueError):
        return 0.0


def esc(s: str) -> str:
    return html.escape(str(s), quote=True)


# ---------------------------------------------------------------------------
# 色の解決（rgb / indexed / theme + tint）
# ---------------------------------------------------------------------------

# Excel の indexed カラー表
INDEXED_COLORS = [
    "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF",
    "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF",
    "800000", "008000", "000080", "808000", "800080", "008080", "C0C0C0", "808080",
    "9999FF", "993366", "FFFFCC", "CCFFFF", "660066", "FF8080", "0066CC", "CCCCFF",
    "000080", "FF00FF", "FFFF00", "00FFFF", "800080", "800000", "008080", "0000FF",
    "00CCFF", "CCFFFF", "CCFFCC", "FFFF99", "99CCFF", "FF99CC", "CC99FF", "FFCC99",
    "3366FF", "33CCCC", "99CC00", "FFCC00", "FF9900", "FF6600", "666699", "969696",
    "003366", "339966", "003300", "333300", "993300", "993366", "333399", "333333",
]

A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"


class ColorResolver:
    """テーマカラー・インデックスカラー・tint を #rrggbb に解決する。"""

    # theme 属性の index → theme1.xml の clrScheme 要素名
    THEME_ORDER = ["lt1", "dk1", "lt2", "dk2",
                   "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
                   "hlink", "folHlink"]

    def __init__(self, theme_xml: bytes | None):
        self.theme: dict[str, str] = {}
        if theme_xml:
            try:
                root = ET.fromstring(theme_xml)
                scheme = root.find(f".//{{{A_NS}}}clrScheme")
                if scheme is not None:
                    for child in scheme:
                        name = child.tag.split("}")[-1]
                        srgb = child.find(f"{{{A_NS}}}srgbClr")
                        sysc = child.find(f"{{{A_NS}}}sysClr")
                        if srgb is not None and srgb.get("val"):
                            self.theme[name] = srgb.get("val")
                        elif sysc is not None:
                            self.theme[name] = sysc.get("lastClr") or "000000"
            except ET.ParseError:
                pass
        # 既定テーマ（Office）
        self.theme.setdefault("dk1", "000000")
        self.theme.setdefault("lt1", "FFFFFF")
        self.theme.setdefault("dk2", "44546A")
        self.theme.setdefault("lt2", "E7E6E6")
        for i, c in enumerate(["4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "70AD47"]):
            self.theme.setdefault(f"accent{i + 1}", c)
        self.theme.setdefault("hlink", "0563C1")
        self.theme.setdefault("folHlink", "954F72")

    def theme_rgb(self, idx: int) -> str:
        try:
            return self.theme.get(self.THEME_ORDER[idx], "000000")
        except IndexError:
            return "000000"

    def resolve(self, color) -> str | None:
        """openpyxl の Color オブジェクト → '#rrggbb'。自動色/未指定は None。"""
        if color is None:
            return None
        ctype = getattr(color, "type", None)
        rgb = None
        if ctype == "rgb":
            v = color.rgb
            if isinstance(v, str) and len(v) >= 6:
                rgb = v[-6:]
        elif ctype == "indexed":
            i = color.indexed
            if isinstance(i, int):
                if i == 64 or i == 65:      # 64=自動(前景) 65=自動(背景)
                    return None
                if 0 <= i < len(INDEXED_COLORS):
                    rgb = INDEXED_COLORS[i]
        elif ctype == "theme":
            t = color.theme
            if isinstance(t, int):
                rgb = self.theme_rgb(t)
        elif ctype == "auto":
            return None
        if rgb is None:
            return None
        tint = getattr(color, "tint", 0.0) or 0.0
        if tint:
            rgb = apply_tint(rgb, tint)
        return "#" + rgb.lower()


def apply_tint(rgb: str, tint: float) -> str:
    """OOXML の tint（明度補正）を適用する。"""
    r, g, b = (int(rgb[i:i + 2], 16) / 255.0 for i in (0, 2, 4))
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    if tint < 0:
        l = l * (1.0 + tint)
    else:
        l = l * (1.0 - tint) + tint
    l = min(1.0, max(0.0, l))
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    return "%02X%02X%02X" % (round(r * 255), round(g * 255), round(b * 255))


def mix(c1: str, c2: str, ratio: float) -> str:
    """#rrggbb 同士を ratio で混色（ratio=1 で c1）。"""
    a = [int(c1[i:i + 2], 16) for i in (1, 3, 5)]
    b = [int(c2[i:i + 2], 16) for i in (1, 3, 5)]
    return "#" + "".join("%02x" % round(x * ratio + y * (1 - ratio)) for x, y in zip(a, b))


# ---------------------------------------------------------------------------
# 表示形式（数値・日付）
# ---------------------------------------------------------------------------

JP_WEEK = ["月", "火", "水", "木", "金", "土", "日"]
ERAS = [  # (開始日, 正式名, 略名, アルファベット)
    (date(2019, 5, 1), "令和", "令", "R"),
    (date(1989, 1, 8), "平成", "平", "H"),
    (date(1926, 12, 25), "昭和", "昭", "S"),
    (date(1912, 7, 30), "大正", "大", "T"),
    (date(1868, 1, 25), "明治", "明", "M"),
]


def _split_sections(fmt: str) -> list[str]:
    out, cur, i, n = [], "", 0, len(fmt)
    in_q = in_br = False
    while i < n:
        c = fmt[i]
        if c == "\\" and i + 1 < n:
            cur += fmt[i:i + 2]
            i += 2
            continue
        if c == '"' and not in_br:
            in_q = not in_q
        elif c == "[" and not in_q:
            in_br = True
        elif c == "]" and in_br:
            in_br = False
        elif c == ";" and not in_q and not in_br:
            out.append(cur)
            cur = ""
            i += 1
            continue
        cur += c
        i += 1
    out.append(cur)
    return out


_COND_RE = re.compile(r"\[(<=|>=|<>|<|>|=)([-+0-9.eE]+)\]")


def _pick_section(sections: list[str], value: float) -> tuple[str, bool]:
    """値に対応するセクションを返す。(書式, 符号を自前で付けるか)"""
    # 条件付きセクション ([>=100] など) を優先評価
    conditional = [(i, s) for i, s in enumerate(sections) if _COND_RE.search(s)]
    if conditional:
        for i, s in conditional:
            m = _COND_RE.search(s)
            op, num = m.group(1), float(m.group(2))
            ok = {"<": value < num, "<=": value <= num, ">": value > num,
                  ">=": value >= num, "=": value == num, "<>": value != num}[op]
            if ok:
                return s, False
        rest = [s for i, s in enumerate(sections) if not _COND_RE.search(s)]
        if rest:
            return rest[0], False
    if len(sections) == 1:
        return sections[0], value < 0
    if value < 0:
        return (sections[1], False) if len(sections) >= 2 and sections[1].strip() else (sections[0], True)
    if value == 0 and len(sections) >= 3:
        return sections[2], False
    return sections[0], False


def _strip_brackets(fmt: str) -> str:
    return re.sub(r"\[(?!h+\]|m+\]|s+\])[^\]]*\]", "", fmt, flags=re.IGNORECASE)


_MASKCHARS = set("#0?.,")


def format_number(value: float, fmt: str) -> str:
    fmt = (fmt or "General").strip()
    if fmt in ("General", "@", "") or fmt.lower() == "general":
        return _general(value)
    sections = _split_sections(fmt)
    sec, own_sign = _pick_section(sections, value)
    sec = _strip_brackets(sec)
    if not sec.strip():
        return ""
    if "E" in sec.upper() and re.search(r"[#0]E[+-]", sec, re.IGNORECASE):
        return _sci(value, sec)
    if re.search(r"[#0?]\s*/\s*[#0?]", sec):     # 分数書式は近似
        return _general(value)

    v = abs(value) if not own_sign else abs(value)
    # トークン分解
    tokens: list[tuple[str, str]] = []
    i, n = 0, len(sec)
    mask = ""
    mask_done = False
    while i < n:
        c = sec[i]
        if c == "\\" and i + 1 < n:
            tokens.append(("lit", sec[i + 1]))
            i += 2
            continue
        if c == '"':
            j = sec.find('"', i + 1)
            j = n if j < 0 else j
            tokens.append(("lit", sec[i + 1:j]))
            i = j + 1
            continue
        if c in "_*":
            tokens.append(("lit", " " if c == "_" else ""))
            i += 2
            continue
        if c == "%":
            tokens.append(("lit", "%"))
            i += 1
            continue
        if c in _MASKCHARS and not mask_done:
            j = i
            while j < n and sec[j] in _MASKCHARS:
                j += 1
            mask = sec[i:j]
            tokens.append(("num", ""))
            mask_done = True
            i = j
            continue
        if c in _MASKCHARS:      # 2つ目以降のマスクは literal 扱い
            tokens.append(("lit", c))
            i += 1
            continue
        tokens.append(("lit", c))
        i += 1

    if "%" in sec.replace('\\%', ""):
        v *= 100

    if not mask:
        return "".join(t[1] for t in tokens)

    int_mask, dot, dec_mask = mask.partition(".")
    trailing = len(int_mask) - len(int_mask.rstrip(","))
    int_mask = int_mask.rstrip(",")
    grouping = "," in int_mask
    int_mask = int_mask.replace(",", "")
    dec_mask = dec_mask.replace(",", "")
    if trailing:
        v /= 1000.0 ** trailing

    decimals = len([c for c in dec_mask if c in "0#?"])
    req_dec = len([c for c in dec_mask if c in "0?"])
    body = f"{v:.{decimals}f}"
    ip, _, dp = body.partition(".")
    if decimals and req_dec < decimals:
        dp = dp[:req_dec] + dp[req_dec:].rstrip("0")
    min_int = int_mask.count("0")
    if len(ip) < min_int:
        ip = ip.rjust(min_int, "0")
    if min_int == 0 and ip == "0" and (dp or dot):
        ip = ""
    if grouping and ip:
        neg = ip.startswith("-")
        digits = ip.lstrip("-")
        digits = "{:,}".format(int(digits)) if digits.isdigit() else digits
        ip = ("-" if neg else "") + digits
    num = ip + (("." + dp) if dp else ("." if (dot and req_dec) else ""))
    if own_sign and value < 0:
        num = "-" + num
    return "".join(num if kind == "num" else txt for kind, txt in tokens)


def _general(value) -> str:
    if isinstance(value, float):
        if value != value or value in (float("inf"), float("-inf")):
            return str(value)
        if float(value).is_integer() and abs(value) < 1e15:
            return str(int(value))
        s = repr(round(value, 10))
        return s
    return str(value)


def _sci(value: float, fmt: str) -> str:
    m = re.search(r"E([+-])0*", fmt, re.IGNORECASE)
    dec = 2
    mm = re.match(r"[#0]*\.?([#0]*)", fmt)
    if mm:
        dec = len(mm.group(1))
    s = f"{value:.{dec}E}"
    if m and m.group(1) == "-":
        s = s.replace("E+", "E")
    return s


_DATE_TOKEN_RE = re.compile(
    r'\[h+\]|\[m+\]|\[s+\]|"[^"]*"|\\.|am/pm|a/p|g+|yyyy|yy|mmmmm|mmmm|mmm|mm|m|'
    r'dddd|ddd|dd|d|aaaa|aaa|hh|h|ss|s|\.0+|e+|.', re.IGNORECASE)

# 上の並びのうち「書式コード」であるもの。区切り記号やリテラルと区別するのに使う
_DATE_FIELD_RE = re.compile(
    r'\[h+\]|\[m+\]|\[s+\]|am/pm|a/p|g+|y+|m+|d+|a{3,4}|h{1,2}|s{1,2}|e+|\.0+',
    re.IGNORECASE)


def _neighbour_code(low: list[str], idx: int, step: int) -> str:
    """区切り記号やリテラルを読み飛ばして、隣にある書式コードを返す。"""
    k = idx + step
    while 0 <= k < len(low):
        if _DATE_FIELD_RE.fullmatch(low[k]):
            return low[k]
        k += step
    return ""


def format_datetime(value, fmt: str) -> str:
    if isinstance(value, timedelta):
        base = None
        total_seconds = value.total_seconds()
        dt = None
    else:
        if isinstance(value, time):
            dt = datetime(1899, 12, 31, value.hour, value.minute, value.second, value.microsecond)
        elif isinstance(value, datetime):
            dt = value
        elif isinstance(value, date):
            dt = datetime(value.year, value.month, value.day)
        else:
            return str(value)
        total_seconds = None

    if not (fmt or "").strip() or fmt.strip().lower() == "general":
        # 表示形式が無い日付。Excel はシリアル値を出すが、読める方が役に立つ
        if dt is None:
            fmt = "[h]:mm:ss"
        elif isinstance(value, time):
            fmt = "h:mm:ss"
        elif dt.hour or dt.minute or dt.second:
            fmt = "yyyy/m/d h:mm:ss"
        else:
            fmt = "yyyy/m/d"

    fmt = _strip_brackets_keep_elapsed(fmt or "")
    sections = _split_sections(fmt)
    fmt = sections[0] if sections else fmt
    toks = _DATE_TOKEN_RE.findall(fmt)
    low = [t.lower() for t in toks]
    ampm = any(t in ("am/pm", "a/p") for t in low)

    out = []
    for idx, tok in enumerate(toks):
        t = tok.lower()
        if t.startswith('"'):
            out.append(tok[1:-1]); continue
        if t.startswith("\\"):
            out.append(tok[1:]); continue
        if dt is None:  # 経過時間のみ
            if t.startswith("[h"):
                out.append(str(int(total_seconds // 3600))); continue
            if t.startswith("[m"):
                out.append(str(int(total_seconds // 60))); continue
            if t.startswith("[s"):
                out.append(str(int(total_seconds))); continue
            if t in ("h", "hh"):
                out.append(f"{int(total_seconds // 3600) % 24:0{len(t)}d}"); continue
            if t in ("m", "mm"):
                out.append(f"{int(total_seconds // 60) % 60:0{len(t)}d}"); continue
            if t in ("s", "ss"):
                out.append(f"{int(total_seconds) % 60:0{len(t)}d}"); continue
            out.append(tok); continue

        if t.startswith("[h") or t.startswith("[m") or t.startswith("[s"):
            secs = (dt - datetime(1899, 12, 30)).total_seconds()
            div = 3600 if t.startswith("[h") else 60 if t.startswith("[m") else 1
            out.append(str(int(secs // div))); continue
        if t in ("yyyy",):
            out.append(f"{dt.year:04d}"); continue
        if t == "yy":
            out.append(f"{dt.year % 100:02d}"); continue
        if t.startswith("g"):
            # g=R / gg=令 / ggg=令和。年の数字は続く e が出すので、ここでは出さない
            era = _era(dt.date())
            out.append(era[3] if len(t) == 1 else era[2] if len(t) == 2 else era[1]); continue
        if t.startswith("e"):
            out.append(f"{_era_year(dt.date()):0{len(t)}d}"); continue
        if t == "aaaa":
            out.append(JP_WEEK[dt.weekday()] + "曜日"); continue
        if t == "aaa":
            out.append(JP_WEEK[dt.weekday()]); continue
        if t == "dddd":
            out.append(dt.strftime("%A")); continue
        if t == "ddd":
            out.append(dt.strftime("%a")); continue
        if t in ("dd", "d"):
            out.append(f"{dt.day:0{len(t)}d}"); continue
        if t in ("mmmmm", "mmmm", "mmm"):
            out.append({"mmmmm": dt.strftime("%b")[0], "mmmm": dt.strftime("%B"),
                        "mmm": dt.strftime("%b")}[t]); continue
        if t in ("m", "mm"):
            # Excel の規則: 時の直後、または秒の直前にある m は「月」ではなく「分」。
            # h:mm の ":" のような区切り記号やリテラルは間にあっても隣とみなす
            prev = _neighbour_code(low, idx, -1)
            nxt = _neighbour_code(low, idx, 1)
            is_minute = (prev in ("h", "hh") or prev.startswith("[h")
                         or nxt in ("s", "ss") or nxt.startswith("[s"))
            v = dt.minute if is_minute else dt.month
            out.append(f"{v:0{len(t)}d}"); continue
        if t in ("h", "hh"):
            hr = dt.hour
            if ampm:
                hr = hr % 12 or 12
            out.append(f"{hr:0{len(t)}d}"); continue
        if t in ("s", "ss"):
            out.append(f"{dt.second:0{len(t)}d}"); continue
        if t.startswith(".0"):
            out.append("." + f"{dt.microsecond:06d}"[:len(t) - 1]); continue
        if t == "am/pm":
            out.append("AM" if dt.hour < 12 else "PM"); continue
        if t == "a/p":
            out.append("A" if dt.hour < 12 else "P"); continue
        out.append(tok)
    return "".join(out)


def _strip_brackets_keep_elapsed(fmt: str) -> str:
    return re.sub(r"\[(?![hms]+\])[^\]]*\]", "", fmt, flags=re.IGNORECASE)


def _era(d: date):
    for start, full, short, alpha in ERAS:
        if d >= start:
            return (start, full, short, alpha)
    return (date(1, 1, 1), "西暦", "西", "A")


def _era_year(d: date) -> int:
    start = _era(d)[0]
    y = d.year - start.year + 1
    return y


def serial_to_datetime(value: float, number_format: str, epoch=None):
    """日付書式なのに数値のまま残っているセルを日付に直す（保険）。

    openpyxl は表示形式を見て日付セルを datetime に変換してくれるが、知らない
    組み込み書式があると素通しになる。そのまま数値として出すと 46255 のような
    シリアル値が見えてしまうので、こちらでも判定して変換する。
    """
    if from_excel is None or not number_format or not is_date_format(number_format):
        return None
    try:
        return from_excel(value, epoch or WINDOWS_EPOCH,
                          timedelta=is_timedelta_format(number_format))
    except (OverflowError, ValueError):
        return None


def format_cell_value(value, number_format: str, epoch=None) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime, date, time, timedelta)):
        return format_datetime(value, number_format)
    if isinstance(value, (int, float)):
        dt = serial_to_datetime(float(value), number_format, epoch)
        if dt is not None:
            return format_datetime(dt, number_format)
        return format_number(float(value), number_format)
    return str(value)


# ---------------------------------------------------------------------------
# 罫線
# ---------------------------------------------------------------------------

BORDER_STYLE_MAP = {
    "thin": (1, "solid"),
    "hair": (1, "solid"),
    "medium": (2, "solid"),
    "thick": (3, "solid"),
    "double": (3, "double"),
    "dotted": (1, "dotted"),
    "dashed": (1, "dashed"),
    "dashDot": (1, "dashed"),
    "dashDotDot": (1, "dotted"),
    "mediumDashed": (2, "dashed"),
    "mediumDashDot": (2, "dashed"),
    "mediumDashDotDot": (2, "dotted"),
    "slantDashDot": (1, "dashed"),
}


def border_css(side, resolver: ColorResolver) -> str | None:
    if side is None or not side.style:
        return None
    w, st = BORDER_STYLE_MAP.get(side.style, (1, "solid"))
    color = resolver.resolve(side.color) or "#000000"
    if side.style == "hair":
        color = mix(color, "#ffffff", 0.55)
    return f"{w}px {st} {color}"


# ---------------------------------------------------------------------------
# 描画オブジェクト（画像・図形）
# ---------------------------------------------------------------------------

XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

MIME = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".gif": "image/gif",
        ".bmp": "image/bmp", ".svg": "image/svg+xml", ".webp": "image/webp",
        ".emf": "image/emf", ".wmf": "image/wmf"}


def _rels_path(part: str) -> str:
    d, b = posixpath.split(part)
    return posixpath.join(d, "_rels", b + ".rels")


def _read_rels(z: zipfile.ZipFile, part: str) -> dict[str, str]:
    p = _rels_path(part)
    try:
        root = ET.fromstring(z.read(p))
    except (KeyError, ET.ParseError):
        return {}
    base = posixpath.dirname(part)
    out = {}
    for rel in root.findall(f"{{{PKG_REL}}}Relationship"):
        tgt = rel.get("Target", "")
        if rel.get("TargetMode") == "External":
            out[rel.get("Id")] = tgt
            continue
        if tgt.startswith("/"):
            out[rel.get("Id")] = tgt.lstrip("/")
        else:
            out[rel.get("Id")] = posixpath.normpath(posixpath.join(base, tgt))
    return out


class DrawingParser:
    """xl/drawings/*.xml を解析して、シートごとの図形リストを返す。"""

    def __init__(self, path: str, resolver: ColorResolver):
        self.resolver = resolver
        self.zip = zipfile.ZipFile(path)
        self.sheet_parts = self._sheet_parts()

    def close(self):
        self.zip.close()

    def _sheet_parts(self) -> dict[str, str]:
        """シート名 → worksheet パート名"""
        try:
            root = ET.fromstring(self.zip.read("xl/workbook.xml"))
        except (KeyError, ET.ParseError):
            return {}
        rels = _read_rels(self.zip, "xl/workbook.xml")
        out = {}
        for sh in root.findall(f"{{{MAIN}}}sheets/{{{MAIN}}}sheet"):
            rid = sh.get(f"{{{REL}}}id")
            if rid and rid in rels:
                out[sh.get("name")] = rels[rid]
        return out

    def shapes_for(self, sheet_name: str, grid) -> list[dict]:
        part = self.sheet_parts.get(sheet_name)
        if not part:
            return []
        try:
            root = ET.fromstring(self.zip.read(part))
        except (KeyError, ET.ParseError):
            return []
        drels = _read_rels(self.zip, part)
        shapes = []
        for tag in (f"{{{MAIN}}}drawing", f"{{{MAIN}}}legacyDrawing"):
            for d in root.findall(tag):
                rid = d.get(f"{{{REL}}}id")
                target = drels.get(rid)
                if not target or "drawings/drawing" not in target:
                    continue
                shapes.extend(self._parse_drawing(target, grid))
        return shapes

    # -- 個々の drawing パート -------------------------------------------------
    def _parse_drawing(self, part: str, grid) -> list[dict]:
        try:
            root = ET.fromstring(self.zip.read(part))
        except (KeyError, ET.ParseError):
            return []
        rels = _read_rels(self.zip, part)
        out = []
        for anchor in root:
            tag = anchor.tag.split("}")[-1]
            if tag not in ("twoCellAnchor", "oneCellAnchor", "absoluteAnchor"):
                continue
            box = self._anchor_box(anchor, tag, grid)
            if box is None:
                continue
            for child in anchor:
                ctag = child.tag.split("}")[-1]
                if ctag in ("from", "to", "ext", "pos", "clientData"):
                    continue
                out.extend(self._emit(child, box, rels, part))
        return out

    def _anchor_box(self, anchor, tag, grid):
        def marker(el):
            if el is None:
                return None
            def g(n, default=0):
                e = el.find(f"{{{XDR}}}{n}")
                return int(e.text) if e is not None and e.text else default
            col, coff, row, roff = g("col"), g("colOff"), g("row"), g("rowOff")
            return grid.x_at(col) + emu_to_px(coff), grid.y_at(row) + emu_to_px(roff)

        if tag == "absoluteAnchor":
            pos = anchor.find(f"{{{XDR}}}pos")
            ext = anchor.find(f"{{{XDR}}}ext")
            if pos is None or ext is None:
                return None
            x, y = emu_to_px(pos.get("x")), emu_to_px(pos.get("y"))
            return x, y, emu_to_px(ext.get("cx")), emu_to_px(ext.get("cy"))
        frm = marker(anchor.find(f"{{{XDR}}}from"))
        if frm is None:
            return None
        if tag == "twoCellAnchor":
            to = marker(anchor.find(f"{{{XDR}}}to"))
            if to is None:
                return None
            return frm[0], frm[1], max(0.0, to[0] - frm[0]), max(0.0, to[1] - frm[1])
        ext = anchor.find(f"{{{XDR}}}ext")
        cx = emu_to_px(ext.get("cx")) if ext is not None else 0
        cy = emu_to_px(ext.get("cy")) if ext is not None else 0
        return frm[0], frm[1], cx, cy

    def _emit(self, node, box, rels, part) -> list[dict]:
        tag = node.tag.split("}")[-1]
        if tag == "pic":
            s = self._picture(node, box, rels)
            return [s] if s else []
        if tag in ("sp", "cxnSp"):
            return [self._shape(node, box)]
        if tag == "grpSp":
            return self._group(node, box, rels, part)
        if tag == "graphicFrame":      # グラフ・SmartArt 等（枠のみ）
            return []
        return []

    def _group(self, node, box, rels, part) -> list[dict]:
        xfrm = node.find(f"{{{XDR}}}grpSpPr/{{{A_NS}}}xfrm")
        out = []
        if xfrm is None:
            for ch in node:
                if ch.tag.split("}")[-1] in ("sp", "pic", "grpSp", "cxnSp"):
                    out.extend(self._emit(ch, box, rels, part))
            return out
        off = xfrm.find(f"{{{A_NS}}}off")
        ext = xfrm.find(f"{{{A_NS}}}ext")
        choff = xfrm.find(f"{{{A_NS}}}chOff")
        chext = xfrm.find(f"{{{A_NS}}}chExt")
        gx, gy, gw, gh = box
        cox = emu_to_px(choff.get("x")) if choff is not None else 0
        coy = emu_to_px(choff.get("y")) if choff is not None else 0
        cw = emu_to_px(chext.get("cx")) if chext is not None else (gw or 1)
        chh = emu_to_px(chext.get("cy")) if chext is not None else (gh or 1)
        sx = (gw / cw) if cw else 1.0
        sy = (gh / chh) if chh else 1.0
        for ch in node:
            ctag = ch.tag.split("}")[-1]
            if ctag not in ("sp", "pic", "grpSp", "cxnSp"):
                continue
            cxf = ch.find(f".//{{{A_NS}}}xfrm")
            if cxf is None:
                out.extend(self._emit(ch, box, rels, part))
                continue
            coff = cxf.find(f"{{{A_NS}}}off")
            cext = cxf.find(f"{{{A_NS}}}ext")
            x = gx + (emu_to_px(coff.get("x")) - cox) * sx if coff is not None else gx
            y = gy + (emu_to_px(coff.get("y")) - coy) * sy if coff is not None else gy
            w = emu_to_px(cext.get("cx")) * sx if cext is not None else gw
            h = emu_to_px(cext.get("cy")) * sy if cext is not None else gh
            out.extend(self._emit(ch, (x, y, w, h), rels, part))
        return out

    def _picture(self, node, box, rels):
        blip = node.find(f".//{{{A_NS}}}blip")
        if blip is None:
            return None
        rid = blip.get(f"{{{REL}}}embed") or blip.get(f"{{{REL}}}link")
        target = rels.get(rid)
        if not target:
            return None
        try:
            data = self.zip.read(target)
        except KeyError:
            return None
        ext = os.path.splitext(target)[1].lower()
        mime = MIME.get(ext, "application/octet-stream")
        if ext in (".emf", ".wmf"):
            return None            # ブラウザで描けない形式は省略
        uri = f"data:{mime};base64,{base64.b64encode(data).decode('ascii')}"
        rot = self._rotation(node)
        name = ""
        nv = node.find(f"{{{XDR}}}nvPicPr/{{{XDR}}}cNvPr")
        if nv is not None:
            name = nv.get("descr") or nv.get("name") or ""
        return {"kind": "img", "box": box, "src": uri, "rot": rot, "alt": name}

    def _rotation(self, node) -> float:
        xfrm = node.find(f".//{{{A_NS}}}xfrm")
        if xfrm is None:
            return 0.0
        rot = xfrm.get("rot")
        return (int(rot) / 60000.0) if rot else 0.0

    def _flips(self, node):
        xfrm = node.find(f".//{{{A_NS}}}xfrm")
        if xfrm is None:
            return False, False
        return xfrm.get("flipH") in ("1", "true"), xfrm.get("flipV") in ("1", "true")

    def _color_of(self, el) -> str | None:
        """a:solidFill 等の色要素 → #rrggbb"""
        if el is None:
            return None
        srgb = el.find(f".//{{{A_NS}}}srgbClr")
        if srgb is not None and srgb.get("val"):
            c = "#" + srgb.get("val").lower()
            return self._modify(c, srgb)
        scheme = el.find(f".//{{{A_NS}}}schemeClr")
        if scheme is not None:
            name = scheme.get("val")
            alias = {"bg1": "lt1", "tx1": "dk1", "bg2": "lt2", "tx2": "dk2",
                     "phClr": "accent1"}.get(name, name)
            rgb = self.resolver.theme.get(alias)
            if rgb:
                return self._modify("#" + rgb.lower(), scheme)
        sysc = el.find(f".//{{{A_NS}}}sysClr")
        if sysc is not None:
            return "#" + (sysc.get("lastClr") or "000000").lower()
        return None

    def _modify(self, color: str, node) -> str:
        """lumMod / lumOff / shade / tint を近似適用。"""
        def val(tag):
            e = node.find(f"{{{A_NS}}}{tag}")
            return int(e.get("val")) / 100000.0 if e is not None and e.get("val") else None
        lum_mod, lum_off = val("lumMod"), val("lumOff")
        shade, tint = val("shade"), val("tint")
        r, g, b = (int(color[i:i + 2], 16) / 255.0 for i in (1, 3, 5))
        h, l, s = colorsys.rgb_to_hls(r, g, b)
        if lum_mod is not None:
            l *= lum_mod
        if lum_off is not None:
            l += lum_off
        if shade is not None:
            l *= shade
        if tint is not None:
            l = l * tint + (1 - tint)
        l = min(1.0, max(0.0, l))
        r, g, b = colorsys.hls_to_rgb(h, l, s)
        return "#%02x%02x%02x" % (round(r * 255), round(g * 255), round(b * 255))

    def _style_color(self, style, tag: str) -> str | None:
        """<a:fillRef>/<a:lnRef> の色。idx="0" は『なし』を意味するので色を返さない。"""
        if style is None:
            return None
        ref = style.find(f"{{{A_NS}}}{tag}")
        if ref is None:
            return None
        try:
            if int(ref.get("idx") or 0) == 0:
                return None            # idx=0 = 塗りなし／線なし
        except ValueError:
            pass
        return self._color_of(ref)

    def _shape(self, node, box) -> dict:
        sp_pr = node.find(f"{{{XDR}}}spPr")
        prst = "rect"
        adj: dict[str, float] = {}
        if sp_pr is not None:
            geom = sp_pr.find(f"{{{A_NS}}}prstGeom")
            if geom is not None:
                prst = geom.get("prst") or "rect"
                for gd in geom.findall(f"{{{A_NS}}}avLst/{{{A_NS}}}gd"):
                    name, fmla = gd.get("name"), gd.get("fmla") or ""
                    m = re.match(r"val\s+(-?\d+)", fmla)
                    if name and m:
                        adj[name] = int(m.group(1)) / 100000.0
            elif sp_pr.find(f"{{{A_NS}}}custGeom") is not None:
                prst = "rect"
        # 塗り
        fill = None
        if sp_pr is not None:
            if sp_pr.find(f"{{{A_NS}}}noFill") is not None:
                fill = None
            else:
                sf = sp_pr.find(f"{{{A_NS}}}solidFill")
                if sf is not None:
                    fill = self._color_of(sf)
                else:
                    gf = sp_pr.find(f"{{{A_NS}}}gradFill")
                    if gf is not None:
                        fill = self._color_of(gf)
        has_explicit_fill = sp_pr is not None and (
            sp_pr.find(f"{{{A_NS}}}noFill") is not None
            or sp_pr.find(f"{{{A_NS}}}solidFill") is not None
            or sp_pr.find(f"{{{A_NS}}}gradFill") is not None)
        # 線
        line_color, line_w, line_dash = None, 1.0, None
        ln = sp_pr.find(f"{{{A_NS}}}ln") if sp_pr is not None else None
        has_explicit_line = False
        head_end = tail_end = None
        if ln is not None:
            has_explicit_line = True
            if ln.get("w"):
                line_w = emu_to_px(ln.get("w"))
            if ln.find(f"{{{A_NS}}}noFill") is not None:
                line_color = None
            else:
                sf = ln.find(f"{{{A_NS}}}solidFill")
                line_color = self._color_of(sf) if sf is not None else None
                if sf is None:
                    has_explicit_line = False
            d = ln.find(f"{{{A_NS}}}prstDash")
            if d is not None:
                line_dash = d.get("val")
            he = ln.find(f"{{{A_NS}}}headEnd")
            te = ln.find(f"{{{A_NS}}}tailEnd")
            head_end = (he.get("type"), he.get("w"), he.get("len")) if he is not None else None
            tail_end = (te.get("type"), te.get("w"), te.get("len")) if te is not None else None
        # スタイル参照（既定色）
        style = node.find(f"{{{XDR}}}style")
        if style is not None:
            if not has_explicit_fill:
                fill = self._style_color(style, "fillRef")
            if not has_explicit_line and line_color is None:
                line_color = self._style_color(style, "lnRef")
        if not has_explicit_fill and fill is None and style is None:
            fill = None
        # 線・コネクタは決して塗らない。スタイル参照の fillRef を拾ってしまうと、
        # 開いたパスが SVG で塗り潰されて巨大な色の塊になる。
        if is_line_shape(prst):
            fill = None
        flip_h, flip_v = self._flips(node)
        return {
            "kind": "shape", "box": box, "prst": prst, "fill": fill, "adj": adj,
            "line": line_color, "line_w": max(line_w, 0.5) if line_color else 0,
            "dash": line_dash, "rot": self._rotation(node),
            "flip_h": flip_h, "flip_v": flip_v,
            "head": head_end, "tail": tail_end,
            "text": self._text_body(node),
        }

    def _text_body(self, node) -> dict | None:
        tx = node.find(f"{{{XDR}}}txBody")
        if tx is None:
            return None
        paras = []
        for p in tx.findall(f"{{{A_NS}}}p"):
            ppr = p.find(f"{{{A_NS}}}pPr")
            algn = ppr.get("algn") if ppr is not None else None
            runs = []
            for r in p:
                t = r.tag.split("}")[-1]
                if t == "br":
                    runs.append({"text": "\n"})
                    continue
                if t not in ("r", "fld"):
                    continue
                tnode = r.find(f"{{{A_NS}}}t")
                if tnode is None or tnode.text is None:
                    continue
                rpr = r.find(f"{{{A_NS}}}rPr")
                run = {"text": tnode.text}
                if rpr is not None:
                    if rpr.get("sz"):
                        run["size"] = int(rpr.get("sz")) / 100.0
                    run["b"] = rpr.get("b") in ("1", "true")
                    run["i"] = rpr.get("i") in ("1", "true")
                    run["u"] = rpr.get("u") not in (None, "none")
                    c = self._color_of(rpr.find(f"{{{A_NS}}}solidFill"))
                    if c:
                        run["color"] = c
                    latin = rpr.find(f"{{{A_NS}}}latin")
                    ea = rpr.find(f"{{{A_NS}}}ea")
                    face = (ea.get("typeface") if ea is not None else None) or \
                           (latin.get("typeface") if latin is not None else None)
                    if face and not face.startswith("+"):
                        run["font"] = face
                runs.append(run)
            if runs:
                paras.append({"algn": algn, "runs": runs})
        if not paras:
            return None
        body = tx.find(f"{{{A_NS}}}bodyPr")
        anchor = body.get("anchor") if body is not None else None
        vert = body.get("vert") if body is not None else None
        return {"paras": paras, "anchor": anchor, "vert": vert}


# 線・コネクタ系（閉じたパスではないので、塗ってはいけない図形）
LINE_PRESETS = {"line", "straightConnector1"}


def is_line_shape(prst: str) -> bool:
    return prst in LINE_PRESETS or prst.startswith(("bentConnector", "curvedConnector"))


def _pts(seq) -> str:
    return "M" + " L".join(f"{x:.2f},{y:.2f}" for x, y in seq) + " Z"


def _conn_path(prst: str, W: float, H: float, adj: dict) -> str | None:
    """カギ線／曲線コネクタ。OOXML の調整値（折れ位置）に従う。"""
    a1 = adj.get("adj1", 0.5)
    a2 = adj.get("adj2", 0.5)
    a3 = adj.get("adj3", 0.5)
    if prst in ("bentConnector2", "curvedConnector2"):
        pts = [(0, 0), (W, 0), (W, H)]
    elif prst in ("bentConnector3", "curvedConnector3"):
        x1 = W * a1
        pts = [(0, 0), (x1, 0), (x1, H), (W, H)]
    elif prst in ("bentConnector4", "curvedConnector4"):
        x1, y2 = W * a1, H * a2
        pts = [(0, 0), (x1, 0), (x1, y2), (W, y2), (W, H)]
    elif prst in ("bentConnector5", "curvedConnector5"):
        x1, y2, x3 = W * a1, H * a2, W * a3
        pts = [(0, 0), (x1, 0), (x1, y2), (x3, y2), (x3, H), (W, H)]
    else:
        return None
    if prst.startswith("bent"):
        return "M" + " L".join(f"{x:.2f},{y:.2f}" for x, y in pts)
    # 曲線コネクタ: 折れ点を二次ベジエの制御点にして角を丸める
    d = [f"M{pts[0][0]:.2f},{pts[0][1]:.2f}"]
    for i in range(1, len(pts) - 1):
        px, py = pts[i - 1]
        cx, cy = pts[i]
        nx, ny = pts[i + 1]
        d.append(f"Q{cx:.2f},{cy:.2f} {(cx+nx)/2:.2f},{(cy+ny)/2:.2f}")
        pts[i] = ((cx + nx) / 2, (cy + ny) / 2)
    d.append(f"L{pts[-1][0]:.2f},{pts[-1][1]:.2f}")
    return " ".join(d)


# 図形 → SVG パス（w,h は 100 正規化ではなく実寸で生成）
def shape_path(prst: str, w: float, h: float, adj: dict | None = None) -> str | None:
    W, H, adj = w, h, adj or {}
    ss = min(W, H)                       # OOXML の "ss"（短辺）
    hc, vc = W / 2, H / 2
    if prst in ("line", "straightConnector1"):
        return f"M0,0 L{W:.2f},{H:.2f}"
    if is_line_shape(prst):
        return _conn_path(prst, W, H, adj)
    if prst in ("triangle", "isoscelesTriangle", "flowChartExtract"):
        return _pts([(W * adj.get("adj", 0.5), 0), (W, H), (0, H)])
    if prst == "rtTriangle":
        return _pts([(0, 0), (0, H), (W, H)])
    if prst in ("diamond", "flowChartDecision"):
        return _pts([(hc, 0), (W, vc), (hc, H), (0, vc)])
    if prst in ("parallelogram", "flowChartInputOutput"):
        d = min(W * adj.get("adj", 0.25), W)
        return _pts([(d, 0), (W, 0), (W - d, H), (0, H)])
    if prst == "trapezoid":
        d = min(W * adj.get("adj", 0.25), hc)
        return _pts([(d, 0), (W - d, 0), (W, H), (0, H)])
    if prst == "hexagon":
        d = min(ss * adj.get("adj", 0.25), hc)
        return _pts([(d, 0), (W - d, 0), (W, vc), (W - d, H), (d, H), (0, vc)])
    if prst == "pentagon":                 # 正五角形（ホームベースは homePlate）
        return _pts([(hc + math.cos(-math.pi / 2 + i * 2 * math.pi / 5) * hc,
                      vc + math.sin(-math.pi / 2 + i * 2 * math.pi / 5) * vc)
                     for i in range(5)])
    if prst == "homePlate":
        d = min(ss * adj.get("adj", 0.16667), W)
        return _pts([(0, 0), (W - d, 0), (W, vc), (W - d, H), (0, H)])
    if prst == "chevron":
        d = min(ss * adj.get("adj", 0.5), W)
        return _pts([(0, 0), (W - d, 0), (W, vc), (W - d, H), (0, H), (d, vc)])
    if prst == "octagon":
        d = min(ss * adj.get("adj", 0.29289), hc, vc)
        return _pts([(d, 0), (W - d, 0), (W, d), (W, H - d), (W - d, H),
                     (d, H), (0, H - d), (0, d)])

    # --- ブロック矢印 -------------------------------------------------------
    # OOXML: adj1 = 軸の太さ（短辺比）, adj2 = 矢尻の長さ（短辺比）
    if prst in ("rightArrow", "leftArrow", "notchedRightArrow", "stripedRightArrow",
                "swooshArrow"):
        dy = ss * min(max(adj.get("adj1", 0.5), 0.0), 1.0) / 2      # 軸の半分の太さ
        dx = min(ss * max(adj.get("adj2", 0.5), 0.0), W)            # 矢尻の長さ
        if prst == "leftArrow":
            p = [(W, vc - dy), (dx, vc - dy), (dx, 0), (0, vc),
                 (dx, H), (dx, vc + dy), (W, vc + dy)]
        else:
            x2 = W - dx
            p = [(0, vc - dy), (x2, vc - dy), (x2, 0), (W, vc),
                 (x2, H), (x2, vc + dy), (0, vc + dy)]
            if prst == "notchedRightArrow":
                p.append((dx * dy / vc if vc else 0, vc))           # 尾のえぐり
        return _pts(p)
    if prst in ("upArrow", "downArrow"):
        dx = ss * min(max(adj.get("adj1", 0.5), 0.0), 1.0) / 2
        dy = min(ss * max(adj.get("adj2", 0.5), 0.0), H)
        if prst == "upArrow":
            return _pts([(hc, 0), (W, dy), (hc + dx, dy), (hc + dx, H),
                         (hc - dx, H), (hc - dx, dy), (0, dy)])
        return _pts([(hc - dx, 0), (hc + dx, 0), (hc + dx, H - dy), (W, H - dy),
                     (hc, H), (0, H - dy), (hc - dx, H - dy)])
    if prst == "leftRightArrow":
        dy = ss * min(max(adj.get("adj1", 0.5), 0.0), 1.0) / 2
        dx = min(ss * max(adj.get("adj2", 0.5), 0.0), hc)
        return _pts([(0, vc), (dx, 0), (dx, vc - dy), (W - dx, vc - dy), (W - dx, 0),
                     (W, vc), (W - dx, H), (W - dx, vc + dy), (dx, vc + dy), (dx, H)])
    if prst == "upDownArrow":
        dx = ss * min(max(adj.get("adj1", 0.5), 0.0), 1.0) / 2
        dy = min(ss * max(adj.get("adj2", 0.5), 0.0), vc)
        return _pts([(hc, 0), (W, dy), (hc + dx, dy), (hc + dx, H - dy), (W, H - dy),
                     (hc, H), (0, H - dy), (hc - dx, H - dy), (hc - dx, dy), (0, dy)])
    if prst == "quadArrow":
        hh = ss * adj.get("adj1", 0.22) / 2       # 軸の半分の太さ
        hd = ss * adj.get("adj2", 0.22)           # 矢尻の半幅
        hl = ss * adj.get("adj3", 0.25)           # 矢尻の長さ
        return _pts([
            (hc, 0), (hc + hd, hl), (hc + hh, hl), (hc + hh, vc - hh),
            (W - hl, vc - hh), (W - hl, vc - hd), (W, vc), (W - hl, vc + hd),
            (W - hl, vc + hh), (hc + hh, vc + hh), (hc + hh, H - hl), (hc + hd, H - hl),
            (hc, H), (hc - hd, H - hl), (hc - hh, H - hl), (hc - hh, vc + hh),
            (hl, vc + hh), (hl, vc + hd), (0, vc), (hl, vc - hd), (hl, vc - hh),
            (hc - hh, vc - hh), (hc - hh, hl), (hc - hd, hl)])
    if prst == "leftRightUpArrow":
        hh = ss * adj.get("adj1", 0.22) / 2
        hd = ss * adj.get("adj2", 0.22)
        hl = ss * adj.get("adj3", 0.25)
        return _pts([
            (hc, 0), (hc + hd, hl), (hc + hh, hl), (hc + hh, H - 2 * hh),
            (W - hl, H - 2 * hh), (W - hl, H - 2 * hh - hd + hh), (W, H - hh),
            (W - hl, H), (W - hl, H - hh + hh), (hl, H), (hl, H - hh + hh),
            (0, H - hh), (hl, H - 2 * hh - hd + hh), (hl, H - 2 * hh),
            (hc - hh, H - 2 * hh), (hc - hh, hl), (hc - hd, hl)])
    if prst in ("bentArrow", "bentUpArrow"):
        th = ss * adj.get("adj1", 0.25)           # 軸の太さ
        hl = ss * adj.get("adj4", 0.4)            # 矢尻の長さ
        hw = th                                    # 矢尻の張り出し
        cx = W - hw - th / 2
        return _pts([(0, H - th), (cx - th / 2, H - th), (cx - th / 2, hl),
                     (cx - hw - th / 2, hl), (cx, 0), (cx + hw + th / 2, hl),
                     (cx + th / 2, hl), (cx + th / 2, H), (0, H)])
    if prst == "uturnArrow":
        th = ss * adj.get("adj1", 0.25)
        hl = ss * adj.get("adj4", 0.4)
        hw = th
        ro = (W - hw) / 2                          # 外側の折り返し半径
        ri = max(ro - th, 0.5)
        ay = min(ro, H - hl)                       # 円弧の中心 y
        return (f"M0,{H:.2f} L0,{ay:.2f} "
                f"A{ro:.2f},{ro:.2f} 0 0 1 {2*ro:.2f},{ay:.2f} "
                f"L{2*ro:.2f},{H-hl:.2f} L{2*ro+hw:.2f},{H-hl:.2f} "
                f"L{2*ro-ri/1:.2f},{H:.2f} L{2*ro-th-hw:.2f},{H-hl:.2f} "
                f"L{2*ro-th:.2f},{H-hl:.2f} L{2*ro-th:.2f},{ay:.2f} "
                f"A{ri:.2f},{ri:.2f} 0 0 0 {th:.2f},{ay:.2f} L{th:.2f},{H:.2f} Z")
    if prst in ("curvedRightArrow", "curvedLeftArrow", "curvedUpArrow", "curvedDownArrow"):
        # 円弧に沿った矢印。縦向きは横向きを作って 90 度入れ替える
        vert = prst in ("curvedUpArrow", "curvedDownArrow")
        A, B = (H, W) if vert else (W, H)
        th = min(A, B) * adj.get("adj1", 0.25)
        hl = min(A, B) * adj.get("adj3", 0.5)
        ro, ri = B / 2, max(B / 2 - th, 0.5)
        xa = max(A - hl, 0.1)
        pt = [(0, ro - th), (xa, ro - th), (xa, ro - th - th / 2), (A, ro),
              (xa, ro + th + th / 2), (xa, ro + th), (0, ro + th)]
        if vert:
            pt = [(y, x) for x, y in pt]
        if prst in ("curvedLeftArrow", "curvedUpArrow"):
            pt = [(W - x, y) if not vert else (x, H - y) for x, y in pt]
        return _pts(pt)

    if prst in ("star4", "star5", "star6", "star8", "star10", "star12", "star16",
                "star24", "star32"):
        n = int(prst[4:])
        inner = {4: 0.16, 5: 0.19, 6: 0.28, 8: 0.37}.get(n, 0.42)
        pts = []
        for i in range(n * 2):
            ang = -math.pi / 2 + i * math.pi / n
            r = 0.5 if i % 2 == 0 else inner
            pts.append((hc + math.cos(ang) * W * r, vc + math.sin(ang) * H * r))
        return _pts(pts)
    return None


# ---------------------------------------------------------------------------
# グリッド（列幅・行高）
# ---------------------------------------------------------------------------

class Grid:
    def __init__(self, ws, max_col: int, max_row: int):
        self.ws = ws
        self.max_col = max_col
        self.max_row = max_row
        fmt = ws.sheet_format
        self.default_col = fmt.defaultColWidth or DEFAULT_COL_WIDTH
        self.default_row_pt = fmt.defaultRowHeight or DEFAULT_ROW_HEIGHT_PT

        widths: dict[int, float] = {}
        hidden_cols: set[int] = set()
        for key, cd in ws.column_dimensions.items():
            lo = getattr(cd, "min", None)
            hi = getattr(cd, "max", None)
            if not lo:
                continue
            hi = hi or lo
            for c in range(lo, min(hi, max_col) + 1):
                if cd.hidden:
                    hidden_cols.add(c)
                    widths[c] = 0.0
                elif cd.width is not None:
                    widths[c] = cd.width
        self.hidden_cols = hidden_cols
        self.col_px = [0] * (max_col + 1)
        for c in range(1, max_col + 1):
            self.col_px[c] = 0 if c in hidden_cols else col_width_to_px(widths.get(c, self.default_col))

        self.hidden_rows: set[int] = set()
        self.row_px = [0.0] * (max_row + 1)
        for r in range(1, max_row + 1):
            rd = ws.row_dimensions.get(r)
            if rd is not None and rd.hidden:
                self.hidden_rows.add(r)
                self.row_px[r] = 0.0
            elif rd is not None and rd.height is not None:
                self.row_px[r] = pt_to_px(rd.height)
            else:
                self.row_px[r] = pt_to_px(self.default_row_pt)

        self._x = [0.0] * (max_col + 2)
        for c in range(1, max_col + 2):
            self._x[c] = self._x[c - 1] + (self.col_px[c - 1] if c - 1 >= 1 else 0)
        self._y = [0.0] * (max_row + 2)
        for r in range(1, max_row + 2):
            self._y[r] = self._y[r - 1] + (self.row_px[r - 1] if r - 1 >= 1 else 0)

    @property
    def width(self) -> float:
        return sum(self.col_px[1:])

    @property
    def height(self) -> float:
        return sum(self.row_px[1:])

    def x_at(self, col0: int) -> float:
        """0 始まり列インデックスの左端 x。使用範囲の外は列定義を見て外挿する。"""
        c = col0 + 1
        if c <= self.max_col + 1:
            return self._x[c]
        x = self._x[self.max_col + 1]
        for cc in range(self.max_col + 1, c):
            cd = self.ws.column_dimensions.get(get_column_letter(cc))
            w = cd.width if (cd is not None and cd.width) else self.default_col
            x += 0 if (cd is not None and cd.hidden) else col_width_to_px(w)
        return x

    def y_at(self, row0: int) -> float:
        r = row0 + 1
        if r <= self.max_row + 1:
            return self._y[r]
        y = self._y[self.max_row + 1]
        for rr in range(self.max_row + 1, r):
            rd = self.ws.row_dimensions.get(rr)
            h = rd.height if (rd is not None and rd.height) else self.default_row_pt
            y += 0 if (rd is not None and rd.hidden) else pt_to_px(h)
        return y


# ---------------------------------------------------------------------------
# セルのスタイル → CSS
# ---------------------------------------------------------------------------

FONT_FALLBACK = ('"Meiryo","Yu Gothic","MS PGothic","Hiragino Kaku Gothic ProN",'
                 '"Noto Sans JP",sans-serif')

H_ALIGN = {"left": "left", "center": "center", "right": "right",
           "justify": "justify", "distributed": "justify",
           "centerContinuous": "center", "fill": "left"}
V_ALIGN = {"top": "top", "center": "middle", "bottom": "bottom",
           "justify": "middle", "distributed": "middle"}


class StyleSheet:
    """同一書式のセルで CSS クラスを共有し、出力サイズを抑える。"""

    def __init__(self, resolver: ColorResolver, gridlines: bool):
        self.resolver = resolver
        self.gridlines = gridlines
        self._by_decl: dict[tuple[str, str], str] = {}
        self._by_style_key: dict[tuple, str] = {}
        self.rules: list[tuple[str, str, str]] = []

    def class_for(self, cell, grid_r: bool = True, grid_b: bool = True,
                  cf: dict | None = None) -> str:
        cfkey = cf_key(cf)
        try:
            key = (tuple(cell._style), grid_r, grid_b, cfkey)
        except Exception:
            key = None
        if key is not None and key in self._by_style_key:
            return self._by_style_key[key]
        td, sp = self._build(cell, grid_r, grid_b, cf)
        name = self._by_decl.get((td, sp))
        if name is None:
            name = f"s{len(self._by_decl)}"
            self._by_decl[(td, sp)] = name
            self.rules.append((name, td, sp))
        if key is not None:
            self._by_style_key[key] = name
        return name

    def _build(self, cell, grid_r: bool = True, grid_b: bool = True,
               cf: dict | None = None) -> tuple[str, str]:
        R = self.resolver
        cf = cf or {}
        font, fill, border, al = cell.font, cell.fill, cell.border, cell.alignment
        td: list[str] = []
        sp: list[str] = []

        # --- 塗り ---
        bg = None
        if fill is not None and fill.fill_type:
            fg = R.resolve(fill.fgColor)
            bgc = R.resolve(fill.bgColor) or "#ffffff"
            if fill.fill_type == "solid":
                bg = fg
            elif fill.fill_type.startswith("gray") or fill.fill_type in (
                    "darkGray", "mediumGray", "lightGray", "gray125", "gray0625"):
                ratio = {"darkGray": .75, "mediumGray": .5, "lightGray": .25,
                         "gray125": .125, "gray0625": .0625}.get(fill.fill_type, .25)
                if fg:
                    bg = mix(fg, bgc, ratio)
            elif fg:
                bg = mix(fg, bgc, 0.5)
        if cf.get("bg"):                       # 条件付き書式の塗りが優先
            bg = cf["bg"]
        if bg:
            td.append(f"background-color:{bg}")

        # --- 罫線 ---
        for side, prop, cfk in (("top", "border-top", "bt"), ("bottom", "border-bottom", "bb"),
                                ("left", "border-left", "bl"), ("right", "border-right", "br")):
            css = cf.get(cfk) or (border_css(getattr(border, side, None), R) if border else None)
            if css:
                td.append(f"{prop}:{css}")
            elif bg:
                # Excel では塗りが目盛線を覆い隠す。罫線が無くても線は出ない。
                continue
            elif self.gridlines and ((prop == "border-right" and grid_r)
                                     or (prop == "border-bottom" and grid_b)):
                # 色を CSS 変数にしておき、画面の切り替えと印刷時の非表示を効かせる。
                # 幅は 1px のまま残るので、消しても方眼の位置はずれない。
                # 隣のセルが罫線を持つ辺には出さない（grid_r / grid_b）。border-collapse は
                # 太さが同じなら左・上側の線を採用するため、目盛線が隣の罫線を消してしまう。
                # --glr / --glb は「はみ出した文字が横切った目盛線」を JS で消すための穴。
                v = "--glr" if prop == "border-right" else "--glb"
                td.append(f"{prop}:1px solid var({v},var(--gl))")
        # 斜線
        if border is not None:
            diag_css = []
            dcolor = R.resolve(border.diagonal.color) if border.diagonal else None
            dcolor = dcolor or "#000000"
            if border.diagonal is not None and border.diagonal.style:
                if border.diagonalDown:
                    diag_css.append(f"linear-gradient(to bottom right,transparent calc(50% - 0.5px),"
                                    f"{dcolor} calc(50% - 0.5px),{dcolor} calc(50% + 0.5px),"
                                    f"transparent calc(50% + 0.5px))")
                if border.diagonalUp:
                    diag_css.append(f"linear-gradient(to top right,transparent calc(50% - 0.5px),"
                                    f"{dcolor} calc(50% - 0.5px),{dcolor} calc(50% + 0.5px),"
                                    f"transparent calc(50% + 0.5px))")
        else:
            diag_css = []
        # データバー（条件付き書式）。斜線より下に敷く
        if cf.get("bar"):
            color, pct, rtl = cf["bar"]
            side = "left" if rtl else "right"
            diag_css.append(f"linear-gradient(to {side},{color} 0 {pct:.1f}%,"
                            f"transparent {pct:.1f}% 100%)")
        if diag_css:
            td.append("background-image:" + ",".join(diag_css))

        # --- フォント ---（条件付き書式の指定があればそちらが勝つ）
        if font is not None:
            if font.name:
                td.append(f'font-family:"{font.name}",{FONT_FALLBACK}')
            if font.sz:
                td.append(f"font-size:{font.sz}pt")
            bold = cf["b"] if "b" in cf else bool(font.b)
            ital = cf["i"] if "i" in cf else bool(font.i)
            if bold:
                td.append("font-weight:700")
            if ital:
                td.append("font-style:italic")
            deco = []
            uline = cf["u"] if "u" in cf else font.u
            if uline:
                deco.append("underline")
            if cf["strike"] if "strike" in cf else font.strike:
                deco.append("line-through")
            if deco:
                td.append("text-decoration:" + " ".join(deco))
                if uline in ("double", "doubleAccounting"):
                    td.append("text-decoration-style:double")
            c = cf.get("color") or R.resolve(font.color)
            if c:
                td.append(f"color:{c}")
            if font.vertAlign == "superscript":
                sp.append("vertical-align:super;font-size:0.7em")
            elif font.vertAlign == "subscript":
                sp.append("vertical-align:sub;font-size:0.7em")

        # --- 配置 ---
        h = H_ALIGN.get(al.horizontal or "", None) if al else None
        v = V_ALIGN.get(al.vertical or "", "bottom") if al else "bottom"
        wrap = bool(al and al.wrapText)
        rot = (al.textRotation if al else 0) or 0
        indent = (al.indent if al else 0) or 0
        transforms: list[str] = []

        td.append(f"vertical-align:{v}")
        if h:
            td.append(f"text-align:{h}")
        if wrap:
            sp.append("white-space:pre-wrap")
            sp.append("position:static;display:block;width:100%")
            if h:
                sp.append(f"text-align:{h}")
            if indent:
                sp.append(f"padding-left:{indent * INDENT_PX}px")
        else:
            sp.append("white-space:pre")
            sp.append("position:absolute")
            # 横方向のアンカー（左右の余白 CELL_PAD は td>i の padding が担う）
            if h == "center":
                sp.append("left:50%")
                transforms.append("translateX(-50%)")
            elif h == "right":
                sp.append(f"right:{indent * INDENT_PX}px")
            else:
                sp.append(f"left:{indent * INDENT_PX}px")
            # 縦方向のアンカー
            if v == "top":
                sp.append("top:0")
            elif v == "middle":
                sp.append("top:50%")
                transforms.append("translateY(-50%)")
            else:
                sp.append("bottom:0")

        if rot:
            if rot == 255:
                sp.append("writing-mode:vertical-rl;text-orientation:upright;"
                          "white-space:pre;letter-spacing:0.05em")
            else:
                deg = -rot if rot <= 90 else (rot - 90)
                transforms.append(f"rotate({deg}deg)")
                sp.append("transform-origin:center center")
        if transforms:
            sp.append("transform:" + " ".join(transforms))
        return ";".join(td), ";".join(sp)

    def css(self) -> str:
        out = []
        for name, td, sp in self.rules:
            if td:
                out.append(f".{name}{{{td}}}")
            if sp:
                out.append(f".{name}>i{{{sp}}}")
        return "\n".join(out)


# ---------------------------------------------------------------------------
# シート → HTML
# ---------------------------------------------------------------------------

def rich_text_html(value) -> str | None:
    """CellRichText（セル内の部分書式）を span 列に変換。"""
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
    except ImportError:
        return None
    if not isinstance(value, CellRichText):
        return None
    parts = []
    for item in value:
        if isinstance(item, str):
            parts.append(esc(item))
            continue
        if isinstance(item, TextBlock):
            f = item.font
            css = []
            if f is not None:
                if f.b:
                    css.append("font-weight:700")
                if f.i:
                    css.append("font-style:italic")
                if f.u:
                    css.append("text-decoration:underline")
                if f.sz:
                    css.append(f"font-size:{f.sz}pt")
                if f.rFont:
                    css.append(f'font-family:"{f.rFont}",{FONT_FALLBACK}')
                col = getattr(f, "color", None)
                if col is not None and getattr(col, "rgb", None) and isinstance(col.rgb, str):
                    css.append("color:#" + col.rgb[-6:].lower())
            parts.append(f'<span style="{";".join(css)}">{esc(str(item.text))}</span>'
                         if css else esc(str(item.text)))
    return "".join(parts)


def load_formulas(src: str, names: list[str]) -> dict[str, dict[tuple[int, int], str]]:
    """数式だけを高速に読み出す（read_only なので書式は読まない）。"""
    try:
        from openpyxl.worksheet.formula import ArrayFormula
    except ImportError:
        ArrayFormula = ()
    out: dict[str, dict[tuple[int, int], str]] = {}
    wb = load_workbook(src, data_only=False, read_only=True)
    try:
        for name in names:
            if name not in wb.sheetnames:
                continue
            m: dict[tuple[int, int], str] = {}
            for row in wb[name].iter_rows():
                for c in row:
                    v = c.value
                    if ArrayFormula and isinstance(v, ArrayFormula):
                        v = v.text
                    if isinstance(v, str) and v.startswith("="):
                        m[(c.row, c.column)] = v
            out[name] = m
    finally:
        wb.close()
    return out


def tidy_formula(f: str) -> str:
    """_xlfn.XLOOKUP のような内部表記を Excel の表示名に戻す。"""
    return re.sub(r"_xlfn\.(_xlws\.)?", "", f)


# ---------------------------------------------------------------------------
# 数式の吹き出し表示（--formula-balloons）
# ---------------------------------------------------------------------------
# 長い数式は括弧と引数区切りの位置で折り、括弧の対応・文字列・数値・参照を
# 色分けして出す。式の中身は変えず、改行とインデントを足すだけ。

FB_WRAP = 50          # 1行に収める目安の文字数（吹き出しの max-width と揃える）
FB_CH_PX = 6.05       # 等幅 10px の 1 文字幅(px)。右端で右寄せに切り替える判定に使う
FB_LINES = 6          # 畳まずに見せる行数。超えた分はホバー／クリックで開く

def _fb_len(s: str) -> int:
    """半角1・全角2で数えた表示幅。折る位置の判断と吹き出しの幅見積もりに使う。"""
    w = 0
    for ch in s:
        o = ord(ch)
        w += 1 if (o < 0x2E80 or 0xFF61 <= o <= 0xFF9F) else 2
    return w


_FB_IDENT = re.compile(r"\$?[^\W\d][\w.$]*")
_FB_NUM = re.compile(r"\d+(?:\.\d*)?(?:[Ee][-+]?\d+)?|\.\d+")
_FB_ERR = re.compile(r"#[A-Za-z0-9_/]*[!?]?")
_FB_A1 = re.compile(r"[A-Za-z]{1,3}\d{1,7}")
# 種別 → CSS クラス。name(名前定義) と op、区切りは色を付けない
_FB_CLS = {"str": "t", "sh": "h", "num": "n", "ref": "r", "fn": "f", "err": "e"}


def _fml_tokens(f: str) -> list[tuple[str, str, int]]:
    """数式を (種別, 文字, 括弧の深さ) に分ける。文字列リテラルの中は触らない。"""
    toks: list[tuple[str, str, int]] = []
    i, n, depth = 0, len(f), 0
    while i < n:
        ch = f[i]
        if ch in "\"'":                     # "文字列" / 'シート名'
            j, q = i + 1, ch
            while j < n:
                if f[j] != q:
                    j += 1
                elif j + 1 < n and f[j + 1] == q:    # "" は文字としての引用符
                    j += 2
                else:
                    j += 1
                    break
            toks.append(("str" if ch == '"' else "sh", f[i:j], depth))
            i = j
            continue
        if ch in "{[":                      # 配列定数・構造参照はまとめて1つ扱い
            j = f.find("}" if ch == "{" else "]", i)
            j = n if j < 0 else j + 1
            toks.append(("num" if ch == "{" else "ref", f[i:j], depth))
            i = j
            continue
        if ch in " \t":
            j = i
            while j < n and f[j] in " \t":
                j += 1
            toks.append(("sp", f[i:j].replace("\t", "  "), depth))
            i = j
            continue
        if ch == "#":                       # #REF! などのエラー値
            m = _FB_ERR.match(f, i)
            toks.append(("err", m.group(0), depth))
            i = m.end()
            continue
        if ch == "(":
            toks.append(("(", ch, depth))
            depth += 1
        elif ch == ")":
            depth = max(0, depth - 1)
            toks.append((")", ch, depth))
        elif ch in ",;":
            toks.append((",", ch, depth))
        elif ch in "\r\n":
            if ch == "\n" or i + 1 >= n or f[i + 1] != "\n":
                toks.append(("nl", "\n", depth))
        else:
            m = _FB_IDENT.match(f, i) or _FB_NUM.match(f, i)
            if m is None:
                toks.append(("op", ch, depth))
            else:
                t, i = m.group(0), m.end()
                if t[0].isdigit() or t[0] == ".":
                    kind = "num"
                elif i < n and f[i] == "(":
                    kind = "fn"
                elif i < n and f[i] == "!":
                    kind = "sh"
                elif _FB_A1.fullmatch(t.replace("$", "")):
                    kind = "ref"
                else:
                    kind = "name"
                toks.append((kind, t, depth))
                continue
        i += 1
    return toks


def format_formula(f: str, width: int = FB_WRAP) -> tuple[str, int, int]:
    """数式を吹き出し向けに整形して HTML を返す。返り値は (HTML, 最長行の文字数, 行数)。

    元から改行が入っている式は書いた人の意図とみなしてその改行を尊重し、
    改行が無くて1行に収まらない式だけを括弧と引数区切りの位置で折る。
    """
    toks = _fml_tokens(tidy_formula(f))
    ind = [-1] * len(toks)          # そのトークンの前で改行するならインデント段数
    if not any(t[0] == "nl" for t in toks):
        pair: dict[int, int] = {}
        stack: list[int] = []
        for i, t in enumerate(toks):
            if t[0] == "(":
                stack.append(i)
            elif t[0] == ")" and stack:
                pair[stack.pop()] = i
        cum = [0]
        for t in toks:
            cum.append(cum[-1] + _fb_len(t[1]))

        def rows_of(lo: int, hi: int, level: int, pts: list[int]) -> list[tuple[int, int]]:
            """候補位置 pts で 1 行に入るところまで詰めて折り、行の範囲を返す。"""
            room = width - level * 2
            starts = [lo]
            k = 0
            while k < len(pts) and cum[hi] - cum[starts[-1]] > room:
                nxt = k
                while nxt < len(pts) and cum[pts[nxt]] - cum[starts[-1]] <= room:
                    nxt += 1
                k = max(nxt - 1, k)
                ind[pts[k]] = level
                starts.append(pts[k])
                k += 1
            starts.append(hi)
            return list(zip(starts, starts[1:]))

        def fold(lo: int, hi: int, level: int) -> None:
            """toks[lo:hi] を level 段のインデントで並べる。収まらなければ折る。

            折り先は引数区切り → 演算子 → 括弧を開く、の順に選ぶ。区切りより
            演算子を先に使うと ">="&$D$2 のような塊が分断されて読みにくい。
            """
            room = width - level * 2
            if cum[hi] - cum[lo] <= room:
                return
            # 直下（括弧の中には入らない）の折れる位置と括弧グループを集める
            seps: list[int] = []
            ops: list[int] = []
            groups: list[tuple[int, int]] = []
            i = lo
            while i < hi:
                kind, text, _ = toks[i]
                j = pair.get(i, -1) if kind == "(" else -1
                if i < j < hi:
                    groups.append((i, j))
                    i = j + 1
                    continue
                if kind == ",":                 # 引数区切りの直後で折る
                    i += 1
                    while i < hi and toks[i][0] == "sp":
                        i += 1
                    if i < hi:
                        seps.append(i)
                    continue
                # 演算子はその手前で折る。単項マイナスや <= の途中では折らない
                if (kind == "op" and text in "&+-*/^<>=" and i > lo
                        and toks[i - 1][0] not in ("op", "(", ",")):
                    ops.append(i)
                i += 1
            for a, b in rows_of(lo, hi, level, seps):
                if cum[b] - cum[a] <= room:
                    continue
                for a2, b2 in rows_of(a, b, level, [p for p in ops if a < p < b]):
                    rest = cum[b2] - cum[a2]
                    if rest <= room:
                        continue
                    # 最後の手段として、中身の長い括弧から順に次の行へ送り出す
                    inner = sorted((g for g in groups if a2 <= g[0] < b2 and g[0] + 1 < g[1]),
                                   key=lambda g: cum[g[1]] - cum[g[0] + 1], reverse=True)
                    for o, c in inner:
                        if rest <= room:
                            break
                        ind[o + 1] = level + 1       # ( の直後で改行
                        ind[c] = level               # ) の直前で改行
                        fold(o + 1, c, level + 1)
                        rest -= cum[c] - cum[o + 1]

        fold(0, len(toks), 0)

    out: list[str] = []
    line = cols = 0
    rows = 1
    for i, (kind, text, depth) in enumerate(toks):
        if ind[i] >= 0:
            out.append("\n" + "  " * ind[i])
            cols, rows, line = max(cols, line), rows + 1, ind[i] * 2
            if kind == "sp":                 # 折った位置の空白は捨てる
                continue
        if kind == "nl":
            cols, rows, line = max(cols, line), rows + 1, 0
            out.append("\n")
            continue
        line += _fb_len(text)
        cls = f"p{depth % 3}" if kind in ("(", ")") else _FB_CLS.get(kind)
        out.append(f'<b class="{cls}">{esc(text)}</b>' if cls else esc(text))
    return "".join(out), max(cols, line), rows


_FB_REF = re.compile(r"(\$?)([A-Za-z]{1,3})(\$?)(\d{1,7})")


def formula_shape(f: str, row: int, col: int) -> str:
    """数式を「そのセルから見た相対位置」に直した形にする。

    相対参照でコピーされた式どうしは同じ文字列になるので、
    「同じ式が縦や横に並んでいる」範囲の判定に使える（R1C1 形式と同じ考え方）。
    """
    out = []
    for kind, text, _ in _fml_tokens(tidy_formula(f)):
        m = _FB_REF.fullmatch(text) if kind == "ref" else None
        if m is None:
            out.append(text)
            continue
        ca, cl, ra, rn = m.groups()
        r, c = int(rn), column_index_from_string(cl.upper())
        out.append((f"R{r}" if ra else f"R[{r - row}]")
                   + (f"C{c}" if ca else f"C[{c - col}]"))
    return "".join(out)


ERROR_TEXTS = {"#VALUE!", "#REF!", "#N/A", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!",
               "#SPILL!", "#CALC!", "#FIELD!", "#BLOCKED!", "#CONNECT!", "#GETTING_DATA"}

_CELL_FILENAME = re.compile(r'CELL\s*\(\s*"filename"', re.I)
_FIND_RBRACKET = re.compile(r'FIND\s*\(\s*"\]"', re.I)
_FIND_LBRACKET = re.compile(r'FIND\s*\(\s*"\["', re.I)
# 「= 他のセルを参照するだけ」の数式。'目次'!$A$1 や Sheet1!A1、A1 を拾う
_PLAIN_REF = re.compile(r"^=\s*(?:(?:'([^']*)'|([^'!\s()+\-*/,]+))!)?\$?([A-Z]{1,3})\$?([1-9]\d*)\s*$")
_STR_LIT = re.compile(r'^"((?:[^"]|"")*)"$')


def _split_concat(expr: str) -> list[str]:
    """文字列リテラルの中を避けて & で分割する。"""
    parts, buf, quoted = [], "", False
    for ch in expr:
        if ch == '"':
            quoted = not quoted
            buf += ch
        elif ch == "&" and not quoted:
            parts.append(buf)
            buf = ""
        else:
            buf += ch
    parts.append(buf)
    return parts


def _eval_refs(formula: str, sheet: str, known: dict[str, dict[tuple[int, int], str]]) -> str | None:
    """解決済みセルへの参照と文字列リテラルの連結だけを評価する。

    ='目次'!$A$1 や ="("&'14.2.用紙'!$A$1&"に記載)" が対象。ひとつでも
    評価できない項があれば None を返し、元の保存値をそのまま使う。
    """
    parts = []
    for tok in _split_concat(formula.lstrip("=")):
        tok = tok.strip()
        lit = _STR_LIT.match(tok)
        if lit:
            parts.append(lit.group(1).replace('""', '"'))
            continue
        ref = _PLAIN_REF.match("=" + tok)
        if not ref:
            return None
        try:
            tgt = (int(ref.group(4)), column_index_from_string(ref.group(3)))
        except ValueError:
            return None
        text = known.get(ref.group(1) or ref.group(2) or sheet, {}).get(tgt)
        if text is None:
            return None
        parts.append(text)
    return "".join(parts)


def resolve_sheet_name_formulas(wb, formulas: dict[str, dict],
                                book_name: str) -> dict[str, dict[tuple[int, int], str]]:
    """CELL("filename") でシート名・ブック名を出す定番の式を解決する。

    CELL は揮発性関数なので Excel は開くたびに再計算するが、保護ビュー
    （「編集を有効にする」が出ている状態）ではパスを取得できず #VALUE! になり、
    その値がそのまま xlsx に保存されていることがある。Excel の画面では見えない
    エラーなので、保存値がエラーのときだけ解決した文字列に差し替える。

    フォルダのパスを取り出す形は、変換した環境のパスが出てしまうので触らない。
    """
    out: dict[str, dict[tuple[int, int], str]] = {}
    for name, m in formulas.items():
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for (r, c), f in m.items():
            if not _CELL_FILENAME.search(f):
                continue
            v = ws.cell(row=r, column=c).value
            if not (isinstance(v, str) and v in ERROR_TEXTS):
                continue
            rb, lb = bool(_FIND_RBRACKET.search(f)), bool(_FIND_LBRACKET.search(f))
            if rb and lb:
                out.setdefault(name, {})[(r, c)] = book_name    # [ ] の中＝ブック名
            elif rb:
                out.setdefault(name, {})[(r, c)] = name         # ] の後ろ＝シート名
    if not out:
        return out

    # 上で解決したセルを参照しているだけのセル（目次など）にも波及させる
    for _ in range(3):
        added = 0
        for name, m in formulas.items():
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            for (r, c), f in m.items():
                if (r, c) in out.get(name, {}):
                    continue
                v = ws.cell(row=r, column=c).value
                if not (isinstance(v, str) and v in ERROR_TEXTS):
                    continue
                text = _eval_refs(f, name, out)
                if text is None:
                    continue
                out.setdefault(name, {})[(r, c)] = text
                added += 1
        if not added:
            break
    return out


# ---------------------------------------------------------------------------
# 数式の簡易評価器（条件付き書式の判定に使う）
#
# 汎用の計算エンジンではない。「$E5="完了"」「MOD(ROW(),2)=1」「AND(...)」のような
# 条件付き書式で実際に使われる式を評価できれば十分なので、評価できない式は
# EvalError にして「ルール不成立」に倒す。
# ---------------------------------------------------------------------------

class _Blank:
    __slots__ = ()

    def __repr__(self):
        return "<blank>"


BLANK = _Blank()
EXCEL_EPOCH = datetime(1899, 12, 30)


class EvalError(Exception):
    """評価できない式。呼び出し側はルール不成立として扱う。"""


_F_TOKEN = re.compile(r"""
    \s*(?:
      (?P<str>"(?:[^"]|"")*")
    | (?P<err>\#(?:VALUE!|REF!|DIV/0!|NAME\?|NULL!|NUM!|N/A|SPILL!|CALC!))
    | (?P<func>[A-Za-z_][A-Za-z0-9_.]*)\s*\(
    | (?P<sheet>(?:'(?:[^']|'')*'|[A-Za-z0-9_.-￿]+)!)
    | (?P<bool>TRUE|FALSE)(?![A-Za-z0-9_])
    | (?P<ref>\$?[A-Za-z]{1,3}\$?[1-9][0-9]*)(?![A-Za-z0-9_(])
    | (?P<name>[A-Za-z_\\][A-Za-z0-9_.\\]*)
    | (?P<num>(?:[0-9]+\.?[0-9]*|\.[0-9]+)(?:[eE][-+]?[0-9]+)?)
    | (?P<op><>|<=|>=|[-+*/^&<>=(),:%])
    )""", re.X)


def _tokenize(s: str) -> list[tuple[str, str]]:
    toks, i, n = [], 0, len(s)
    while i < n:
        if s[i].isspace():
            i += 1
            continue
        m = _F_TOKEN.match(s, i)
        if not m or m.end() == i:
            raise EvalError(f"読めない文字: {s[i]!r}")
        i = m.end()
        toks.append((m.lastgroup, m.group(m.lastgroup)))
    return toks


def _serial(v) -> float:
    if isinstance(v, datetime):
        return (v - EXCEL_EPOCH).total_seconds() / 86400.0
    if isinstance(v, date):
        return float((datetime(v.year, v.month, v.day) - EXCEL_EPOCH).days)
    if isinstance(v, time):
        return (v.hour * 3600 + v.minute * 60 + v.second) / 86400.0
    if isinstance(v, timedelta):
        return v.total_seconds() / 86400.0
    raise EvalError("日付として扱えない")


def _num(v) -> float:
    if v is BLANK or v is None:
        return 0.0
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, (datetime, date, time, timedelta)):
        return _serial(v)
    if isinstance(v, str):
        if v in ERROR_TEXTS:
            raise EvalError(v)
        try:
            return float(v.strip())
        except ValueError:
            raise EvalError("数値にできない")
    raise EvalError("数値にできない")


def _text(v) -> str:
    if v is BLANK or v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def _rank(v) -> int:
    """Excel の型順: 数値 < 文字列 < FALSE < TRUE"""
    if isinstance(v, bool):
        return 3 if v else 2
    if isinstance(v, str):
        return 1
    return 0


def _compare(a, b) -> int:
    if a is BLANK and b is BLANK:
        return 0
    if a is BLANK:
        a = "" if isinstance(b, str) else (False if isinstance(b, bool) else 0)
    if b is BLANK:
        b = "" if isinstance(a, str) else (False if isinstance(a, bool) else 0)
    ra, rb = _rank(a), _rank(b)
    if ra != rb:
        return -1 if ra < rb else 1
    if ra == 1:
        x, y = a.upper(), b.upper()
        return (x > y) - (x < y)
    if ra >= 2:
        return 0
    x, y = _num(a), _num(b)
    return (x > y) - (x < y)


def _truthy(v) -> bool:
    if v is BLANK or v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        if v.upper() == "TRUE":
            return True
        if v.upper() == "FALSE" or v == "":
            return False
        raise EvalError("真偽にできない")
    return _num(v) != 0


class EvalCtx:
    """式を評価するときのシート・現在セル・相対参照のずれ。"""

    def __init__(self, wb, ws, row: int, col: int, anchor: tuple[int, int], today=None):
        self.wb, self.ws = wb, ws
        self.row, self.col = row, col
        self.dr = row - anchor[0]
        self.dc = col - anchor[1]
        self.today = today or date.today()

    def cell(self, sheet: str | None, col: int, row: int):
        ws = self.ws
        if sheet:
            if self.wb is None or sheet not in self.wb.sheetnames:
                raise EvalError(f"シートが無い: {sheet}")
            ws = self.wb[sheet]
        if row < 1 or col < 1 or row > 1048576 or col > 16384:
            raise EvalError("参照範囲外")
        v = ws.cell(row=row, column=col).value
        return BLANK if v is None else v


_REF_RE = re.compile(r"^(\$?)([A-Za-z]{1,3})(\$?)([1-9][0-9]*)$")


class _Range:
    __slots__ = ("values",)

    def __init__(self, values):
        self.values = values


class _Parser:
    def __init__(self, toks, ctx: EvalCtx):
        self.t, self.i, self.ctx = toks, 0, ctx

    def peek(self):
        return self.t[self.i] if self.i < len(self.t) else (None, None)

    def take(self):
        tok = self.peek()
        self.i += 1
        return tok

    def eat_op(self, *ops) -> str | None:
        k, v = self.peek()
        if k == "op" and v in ops:
            self.i += 1
            return v
        return None

    # 比較 < 連結 < 加減 < 乗除 < べき乗 < 単項 < 一次
    def parse(self):
        v = self.compare()
        if self.i != len(self.t):
            raise EvalError("式の末尾が余っている")
        return v

    def compare(self):
        left = self.concat()
        while True:
            op = self.eat_op("=", "<>", "<", ">", "<=", ">=")
            if not op:
                return left
            right = self.concat()
            c = _compare(_scalar(left), _scalar(right))
            left = {"=": c == 0, "<>": c != 0, "<": c < 0, ">": c > 0,
                    "<=": c <= 0, ">=": c >= 0}[op]

    def concat(self):
        left = self.addsub()
        while self.eat_op("&"):
            left = _text(_scalar(left)) + _text(_scalar(self.addsub()))
        return left

    def addsub(self):
        left = self.muldiv()
        while True:
            op = self.eat_op("+", "-")
            if not op:
                return left
            r = self.muldiv()
            left = _num(_scalar(left)) + _num(_scalar(r)) if op == "+" \
                else _num(_scalar(left)) - _num(_scalar(r))

    def muldiv(self):
        left = self.power()
        while True:
            op = self.eat_op("*", "/")
            if not op:
                return left
            r = _num(_scalar(self.power()))
            if op == "/":
                if r == 0:
                    raise EvalError("#DIV/0!")
                left = _num(_scalar(left)) / r
            else:
                left = _num(_scalar(left)) * r

    def power(self):
        left = self.unary()
        if self.eat_op("^"):
            return _num(_scalar(left)) ** _num(_scalar(self.power()))
        return left

    def unary(self):
        op = self.eat_op("-", "+")
        if op == "-":
            return -_num(_scalar(self.unary()))
        if op == "+":
            return _num(_scalar(self.unary()))
        return self.postfix()

    def postfix(self):
        v = self.primary()
        while self.eat_op("%"):
            v = _num(_scalar(v)) / 100.0
        return v

    def primary(self):
        kind, val = self.take()
        if kind == "num":
            return float(val)
        if kind == "str":
            return val[1:-1].replace('""', '"')
        if kind == "bool":
            return val.upper() == "TRUE"
        if kind == "err":
            raise EvalError(val)
        if kind == "op" and val == "(":
            v = self.compare()
            if not self.eat_op(")"):
                raise EvalError("')' が無い")
            return v
        if kind == "func":
            return self.call(val.upper())
        if kind in ("sheet", "ref"):
            return self.reference(kind, val)
        if kind == "name":
            raise EvalError(f"未対応の名前: {val}")
        raise EvalError("式が読めない")

    def reference(self, kind, val):
        sheet = None
        if kind == "sheet":
            sheet = val[:-1]
            if sheet.startswith("'"):
                sheet = sheet[1:-1].replace("''", "'")
            k2, v2 = self.take()
            if k2 != "ref":
                raise EvalError("シート参照の後がセルでない")
            val = v2
        a = self._addr(val)
        if self.eat_op(":"):
            k2, v2 = self.take()
            if k2 != "ref":
                raise EvalError("範囲の終わりがセルでない")
            b = self._addr(v2)
            r1, r2 = sorted((a[1], b[1]))
            c1, c2 = sorted((a[0], b[0]))
            if (r2 - r1 + 1) * (c2 - c1 + 1) > 200000:
                raise EvalError("範囲が大きすぎる")
            return _Range([self.ctx.cell(sheet, c, r)
                           for r in range(r1, r2 + 1) for c in range(c1, c2 + 1)])
        return self.ctx.cell(sheet, a[0], a[1])

    def _addr(self, ref: str) -> tuple[int, int]:
        m = _REF_RE.match(ref)
        if not m:
            raise EvalError(f"参照が読めない: {ref}")
        col = column_index_from_string(m.group(2).upper())
        row = int(m.group(4))
        if not m.group(1):
            col += self.ctx.dc
        if not m.group(3):
            row += self.ctx.dr
        return col, row

    def args(self) -> list:
        out = []
        if self.eat_op(")"):
            return out
        while True:
            out.append(self.compare())
            if self.eat_op(","):
                continue
            if self.eat_op(")"):
                return out
            raise EvalError("引数の区切りが読めない")

    def call(self, name: str):
        args = self.args()
        fn = _FUNCS.get(name)
        if fn is None:
            raise EvalError(f"未対応の関数: {name}")
        return fn(self.ctx, args)


def _scalar(v):
    if isinstance(v, _Range):
        if len(v.values) != 1:
            raise EvalError("単一セルが必要")
        return v.values[0]
    return v


def _flat(args) -> list:
    out = []
    for a in args:
        out.extend(a.values if isinstance(a, _Range) else [a])
    return out


def _nums(args) -> list[float]:
    out = []
    for v in _flat(args):
        if v is BLANK or v is None or isinstance(v, str) or isinstance(v, bool):
            continue
        out.append(_num(v))
    return out


def _criteria(crit):
    """COUNTIF などの条件文字列 → 判定関数。"""
    if isinstance(crit, str):
        m = re.match(r"^\s*(<>|>=|<=|=|<|>)\s*(.*)$", crit)
        if m:
            op, rest = m.group(1), m.group(2)
            try:
                target = float(rest)
            except ValueError:
                target = rest
            return lambda v: {"=": lambda c: c == 0, "<>": lambda c: c != 0,
                              "<": lambda c: c < 0, ">": lambda c: c > 0,
                              "<=": lambda c: c <= 0, ">=": lambda c: c >= 0}[op](
                _compare(v, target))
    return lambda v: _compare(v, crit) == 0


def _f_if(ctx, a):
    if len(a) < 2:
        raise EvalError("IF の引数不足")
    return a[1] if _truthy(_scalar(a[0])) else (a[2] if len(a) > 2 else False)


def _f_iferror(ctx, a):
    try:
        return _scalar(a[0])
    except EvalError:
        return a[1] if len(a) > 1 else ""


def _f_weekday(ctx, a):
    d = _num(_scalar(a[0]))
    typ = int(_num(_scalar(a[1]))) if len(a) > 1 else 1
    dow = int(d) % 7          # 1900/1/1(=1) は日曜
    if typ == 1:
        return float(dow if dow else 7)
    if typ == 2:
        return float(dow - 1 if dow >= 1 else 7)
    if typ == 3:
        return float((dow - 2) % 7)
    return float(dow if dow else 7)


def _dt(v):
    n = _num(v)
    return EXCEL_EPOCH + timedelta(days=n)


_FUNCS = {
    "AND": lambda c, a: all(_truthy(v) for v in _flat(a)),
    "OR": lambda c, a: any(_truthy(v) for v in _flat(a)),
    "NOT": lambda c, a: not _truthy(_scalar(a[0])),
    "TRUE": lambda c, a: True,
    "FALSE": lambda c, a: False,
    "IF": _f_if,
    "IFERROR": _f_iferror,
    "IFNA": _f_iferror,
    "ISBLANK": lambda c, a: _scalar(a[0]) is BLANK,
    "ISNUMBER": lambda c, a: isinstance(_scalar(a[0]), (int, float, datetime, date, time))
                             and not isinstance(_scalar(a[0]), bool),
    "ISTEXT": lambda c, a: isinstance(_scalar(a[0]), str)
                           and _scalar(a[0]) not in ERROR_TEXTS,
    "ISERROR": lambda c, a: isinstance(_scalar(a[0]), str) and _scalar(a[0]) in ERROR_TEXTS,
    "ISERR": lambda c, a: isinstance(_scalar(a[0]), str) and _scalar(a[0]) in ERROR_TEXTS,
    "ISNA": lambda c, a: _scalar(a[0]) == "#N/A",
    "ISEVEN": lambda c, a: int(_num(_scalar(a[0]))) % 2 == 0,
    "ISODD": lambda c, a: int(_num(_scalar(a[0]))) % 2 == 1,
    "N": lambda c, a: _num(_scalar(a[0])),
    "T": lambda c, a: _scalar(a[0]) if isinstance(_scalar(a[0]), str) else "",
    "LEN": lambda c, a: float(len(_text(_scalar(a[0])))),
    "TRIM": lambda c, a: " ".join(_text(_scalar(a[0])).split()),
    "UPPER": lambda c, a: _text(_scalar(a[0])).upper(),
    "LOWER": lambda c, a: _text(_scalar(a[0])).lower(),
    "LEFT": lambda c, a: _text(_scalar(a[0]))[:int(_num(_scalar(a[1]))) if len(a) > 1 else 1],
    "RIGHT": lambda c, a: _text(_scalar(a[0]))[-(int(_num(_scalar(a[1]))) if len(a) > 1 else 1):]
                          if (int(_num(_scalar(a[1]))) if len(a) > 1 else 1) else "",
    "MID": lambda c, a: _text(_scalar(a[0]))[max(int(_num(_scalar(a[1]))) - 1, 0):
                                             max(int(_num(_scalar(a[1]))) - 1, 0)
                                             + int(_num(_scalar(a[2])))],
    "EXACT": lambda c, a: _text(_scalar(a[0])) == _text(_scalar(a[1])),
    "CONCATENATE": lambda c, a: "".join(_text(v) for v in _flat(a)),
    "VALUE": lambda c, a: _num(_scalar(a[0])),
    "ABS": lambda c, a: abs(_num(_scalar(a[0]))),
    "INT": lambda c, a: float(math.floor(_num(_scalar(a[0])))),
    "MOD": lambda c, a: _num(_scalar(a[0])) - _num(_scalar(a[1]))
                        * math.floor(_num(_scalar(a[0])) / _num(_scalar(a[1]))),
    "ROUND": lambda c, a: round(_num(_scalar(a[0])), int(_num(_scalar(a[1]))) if len(a) > 1 else 0),
    "ROW": lambda c, a: float(c.row if not a else _row_of(a[0], c)),
    "COLUMN": lambda c, a: float(c.col if not a else _col_of(a[0], c)),
    "SUM": lambda c, a: float(sum(_nums(a))),
    "COUNT": lambda c, a: float(len(_nums(a))),
    "COUNTA": lambda c, a: float(sum(1 for v in _flat(a) if v is not BLANK and v != "")),
    "COUNTBLANK": lambda c, a: float(sum(1 for v in _flat(a) if v is BLANK or v == "")),
    "AVERAGE": lambda c, a: (float(sum(_nums(a))) / len(_nums(a))) if _nums(a)
                            else _raise("#DIV/0!"),
    "MAX": lambda c, a: max(_nums(a)) if _nums(a) else 0.0,
    "MIN": lambda c, a: min(_nums(a)) if _nums(a) else 0.0,
    "COUNTIF": lambda c, a: float(sum(1 for v in _flat([a[0]])
                                      if _criteria(_scalar(a[1]))(v))),
    "SUMIF": lambda c, a: float(sum(_num(v) for v in _flat([a[0]])
                                    if _criteria(_scalar(a[1]))(v) and not isinstance(v, str))),
    "TODAY": lambda c, a: float((datetime(c.today.year, c.today.month, c.today.day)
                                 - EXCEL_EPOCH).days),
    "NOW": lambda c, a: float((datetime(c.today.year, c.today.month, c.today.day)
                               - EXCEL_EPOCH).days),
    "YEAR": lambda c, a: float(_dt(_scalar(a[0])).year),
    "MONTH": lambda c, a: float(_dt(_scalar(a[0])).month),
    "DAY": lambda c, a: float(_dt(_scalar(a[0])).day),
    "DATE": lambda c, a: float((datetime(int(_num(_scalar(a[0]))), int(_num(_scalar(a[1]))),
                                         int(_num(_scalar(a[2])))) - EXCEL_EPOCH).days),
    "WEEKDAY": _f_weekday,
    "SEARCH": lambda c, a: float(_text(_scalar(a[1])).upper()
                                 .index(_text(_scalar(a[0])).upper()) + 1)
                           if _text(_scalar(a[0])).upper() in _text(_scalar(a[1])).upper()
                           else _raise("#VALUE!"),
    "FIND": lambda c, a: float(_text(_scalar(a[1])).index(_text(_scalar(a[0]))) + 1)
                         if _text(_scalar(a[0])) in _text(_scalar(a[1])) else _raise("#VALUE!"),
}


def _raise(msg):
    raise EvalError(msg)


def _row_of(arg, ctx):
    raise EvalError("ROW(参照) は未対応")


def _col_of(arg, ctx):
    raise EvalError("COLUMN(参照) は未対応")


_TOKEN_CACHE: dict[str, list | str] = {}


def eval_formula(expr: str, ctx: EvalCtx):
    """式を評価する。評価できないときは EvalError。

    同じ式が範囲内の全セルで評価されるので、字句解析の結果は使い回す。"""
    s = expr.strip()
    if s.startswith("="):
        s = s[1:]
    if not s:
        raise EvalError("空の式")
    toks = _TOKEN_CACHE.get(s)
    if toks is None:
        try:
            toks = _tokenize(s)
        except EvalError as e:
            _TOKEN_CACHE[s] = str(e)
            raise
        if len(_TOKEN_CACHE) < 5000:
            _TOKEN_CACHE[s] = toks
    elif isinstance(toks, str):
        raise EvalError(toks)
    return _scalar(_Parser(toks, ctx).parse())


# ---------------------------------------------------------------------------
# 条件付き書式
# ---------------------------------------------------------------------------

def cf_key(cf: dict | None):
    """条件付き書式の結果を CSS クラスのキーに使えるタプルにする。"""
    if not cf:
        return None
    return tuple(sorted((k, v) for k, v in cf.items() if k not in ("icon", "novalue")))


# アイコンセット: (色, 形) の並び。index 0 が「最も低い」
_ICON_SETS = {
    "3Arrows": [("#d62f2f", "down"), ("#e8b400", "right"), ("#4f9e37", "up")],
    "3ArrowsGray": [("#888888", "down"), ("#888888", "right"), ("#888888", "up")],
    "4Arrows": [("#d62f2f", "down"), ("#e08c00", "downright"), ("#c9b400", "upright"),
                ("#4f9e37", "up")],
    "4ArrowsGray": [("#888888", "down"), ("#888888", "downright"), ("#888888", "upright"),
                    ("#888888", "up")],
    "5Arrows": [("#d62f2f", "down"), ("#e08c00", "downright"), ("#e8b400", "right"),
                ("#a3b800", "upright"), ("#4f9e37", "up")],
    "5ArrowsGray": [("#888888", "down"), ("#888888", "downright"), ("#888888", "right"),
                    ("#888888", "upright"), ("#888888", "up")],
    "3TrafficLights1": [("#d62f2f", "circle"), ("#e8b400", "circle"), ("#4f9e37", "circle")],
    "3TrafficLights2": [("#d62f2f", "circle"), ("#e8b400", "circle"), ("#4f9e37", "circle")],
    "4TrafficLights": [("#3b3b3b", "circle"), ("#d62f2f", "circle"), ("#e8b400", "circle"),
                       ("#4f9e37", "circle")],
    "3Signs": [("#d62f2f", "diamond"), ("#e8b400", "triangle"), ("#4f9e37", "circle")],
    "3Symbols": [("#d62f2f", "cross"), ("#e8b400", "excl"), ("#4f9e37", "check")],
    "3Symbols2": [("#d62f2f", "cross"), ("#e8b400", "excl"), ("#4f9e37", "check")],
    "3Flags": [("#d62f2f", "flag"), ("#e8b400", "flag"), ("#4f9e37", "flag")],
    "4RedToBlack": [("#3b3b3b", "circle"), ("#6b3b3b", "circle"), ("#c46a5a", "circle"),
                    ("#d62f2f", "circle")],
    "3Stars": [("#c9a227", "star0"), ("#c9a227", "star1"), ("#c9a227", "star2")],
    "5Quarters": [("#5b5b5b", "q0"), ("#5b5b5b", "q1"), ("#5b5b5b", "q2"),
                  ("#5b5b5b", "q3"), ("#5b5b5b", "q4")],
    "4Rating": [("#4c78a8", "r1"), ("#4c78a8", "r2"), ("#4c78a8", "r3"), ("#4c78a8", "r4")],
    "5Rating": [("#4c78a8", "r1"), ("#4c78a8", "r2"), ("#4c78a8", "r3"), ("#4c78a8", "r4"),
                ("#4c78a8", "r5")],
    "5Boxes": [("#4c78a8", "b1"), ("#4c78a8", "b2"), ("#4c78a8", "b3"), ("#4c78a8", "b4"),
               ("#4c78a8", "b5")],
    "3TrafficLights": [("#d62f2f", "circle"), ("#e8b400", "circle"), ("#4f9e37", "circle")],
}

_ARROW_ANGLE = {"up": -90, "upright": -45, "right": 0, "downright": 45, "down": 90}


def icon_svg(setname: str, idx: int) -> str:
    """アイコンセットの1個を 1em 角のインライン SVG にする。"""
    icons = _ICON_SETS.get(setname) or _ICON_SETS["3TrafficLights1"]
    color, shape = icons[min(max(idx, 0), len(icons) - 1)]
    b = ""
    if shape in _ARROW_ANGLE:
        ang = _ARROW_ANGLE[shape]
        b = (f'<g transform="rotate({ang} 8 8)">'
             f'<path d="M2,6 L9,6 L9,2.5 L14.5,8 L9,13.5 L9,10 L2,10 Z" fill="{color}"/></g>')
    elif shape == "circle":
        b = f'<circle cx="8" cy="8" r="6.2" fill="{color}"/>'
    elif shape == "diamond":
        b = f'<path d="M8,1.4 L14.6,8 L8,14.6 L1.4,8 Z" fill="{color}"/>'
    elif shape == "triangle":
        b = f'<path d="M8,1.6 L15,14 L1,14 Z" fill="{color}"/>'
    elif shape == "check":
        b = (f'<path d="M2.5,8.5 L6.3,12.4 L13.6,3.6" fill="none" stroke="{color}" '
             f'stroke-width="2.6" stroke-linecap="round" stroke-linejoin="round"/>')
    elif shape == "cross":
        b = (f'<path d="M3,3 L13,13 M13,3 L3,13" fill="none" stroke="{color}" '
             f'stroke-width="2.6" stroke-linecap="round"/>')
    elif shape == "excl":
        b = (f'<path d="M8,2 L8,10" stroke="{color}" stroke-width="2.6" stroke-linecap="round"/>'
             f'<circle cx="8" cy="13.4" r="1.5" fill="{color}"/>')
    elif shape == "flag":
        b = (f'<path d="M3.2,1.5 L3.2,14.5" stroke="{color}" stroke-width="1.8"/>'
             f'<path d="M4.4,2.2 L13.5,2.2 L11,5.6 L13.5,9 L4.4,9 Z" fill="{color}"/>')
    elif shape.startswith("star"):
        n = int(shape[4:])
        pts = []
        for i in range(10):
            ang = -math.pi / 2 + i * math.pi / 5
            r = 7.0 if i % 2 == 0 else 2.9
            pts.append(f"{8 + math.cos(ang)*r:.1f},{8 + math.sin(ang)*r:.1f}")
        star = "M" + " L".join(pts) + " Z"
        fillpct = [0, 50, 100][min(n, 2)]
        b = (f'<defs><linearGradient id="g{setname}{idx}"><stop offset="{fillpct}%" '
             f'stop-color="{color}"/><stop offset="{fillpct}%" stop-color="#ffffff"/>'
             f'</linearGradient></defs>'
             f'<path d="{star}" fill="url(#g{setname}{idx})" stroke="{color}" '
             f'stroke-width="1"/>')
    elif shape.startswith("q"):
        n = int(shape[1:])
        b = f'<circle cx="8" cy="8" r="6.4" fill="none" stroke="{color}" stroke-width="1.4"/>'
        if n:
            ang = -math.pi / 2 + n * math.pi / 2
            large = 1 if n > 2 else 0
            ex, ey = 8 + math.cos(ang) * 6.4, 8 + math.sin(ang) * 6.4
            if n == 4:
                b += f'<circle cx="8" cy="8" r="6.4" fill="{color}"/>'
            else:
                b += (f'<path d="M8,8 L8,1.6 A6.4,6.4 0 {large} 1 {ex:.2f},{ey:.2f} Z" '
                      f'fill="{color}"/>')
    elif shape.startswith(("r", "b")):
        n = int(shape[1:])
        total = 4 if setname == "4Rating" else 5
        b = ""
        for i in range(total):
            fill = color if i < n else "#ffffff"
            hgt = 3 + (i + 1) * (10.0 / total)
            b += (f'<rect x="{0.6 + i*(15.0/total):.2f}" y="{15-hgt:.2f}" '
                  f'width="{13.0/total:.2f}" height="{hgt:.2f}" fill="{fill}" '
                  f'stroke="{color}" stroke-width="0.7"/>')
    return (f'<svg class="cfi" viewBox="0 0 16 16" width="1em" height="1em" '
            f'aria-hidden="true">{b}</svg>')


class CondFormat:
    """シートの条件付き書式を評価して、セル → 追加書式 の対応表を作る。"""

    MAX_CELLS = 400_000

    def __init__(self, wb, ws, resolver: ColorResolver, bounds, today=None):
        self.wb, self.ws, self.R = wb, ws, resolver
        self.today = today or date.today()
        self.map: dict[tuple[int, int], dict] = {}
        self.n_rules = 0
        self.truncated = False
        self.unsupported: set[str] = set()
        self.uneval: set[str] = set()      # 評価できなかった数式ルール（理由つき）
        self._build(bounds)

    # -- 準備 ---------------------------------------------------------------
    def _build(self, bounds):
        min_row, min_col, max_row, max_col = bounds
        entries = []
        for cf in self.ws.conditional_formatting:
            rngs, anchor_r, anchor_c, full = [], None, None, []
            for rr in cf.sqref.ranges:
                r1, c1 = rr.min_row or 1, rr.min_col or 1
                r2, c2 = rr.max_row or r1, rr.max_col or c1
                anchor_r = r1 if anchor_r is None else min(anchor_r, r1)
                anchor_c = c1 if anchor_c is None else min(anchor_c, c1)
                full.append((r1, c1, min(r2, self.ws.max_row or r2),
                             min(c2, self.ws.max_column or c2)))
                a1, b1 = max(r1, min_row), max(c1, min_col)
                a2, b2 = min(r2, max_row), min(c2, max_col)
                if a1 <= a2 and b1 <= b2:
                    rngs.append((a1, b1, a2, b2))
            if not rngs:
                continue
            for rule in cf.rules:
                entries.append({
                    "rule": rule, "ranges": rngs, "full": full,
                    "anchor": (anchor_r or 1, anchor_c or 1),
                    "prio": rule.priority if rule.priority is not None else 10 ** 6,
                    "stop": bool(rule.stopIfTrue),
                    "agg": None,
                })
        if not entries:
            return
        entries.sort(key=lambda e: e["prio"])
        self.n_rules = len(entries)

        cell_rules: dict[tuple[int, int], list[int]] = {}
        budget = self.MAX_CELLS
        for idx, ent in enumerate(entries):
            for (r1, c1, r2, c2) in ent["ranges"]:
                n = (r2 - r1 + 1) * (c2 - c1 + 1)
                if n > budget:
                    self.truncated = True
                    break
                budget -= n
                for r in range(r1, r2 + 1):
                    for c in range(c1, c2 + 1):
                        cell_rules.setdefault((r, c), []).append(idx)

        for (r, c), idxs in cell_rules.items():
            acc: dict = {}
            for idx in idxs:
                ent = entries[idx]
                try:
                    res = self._apply(ent, r, c)
                except EvalError as e:
                    # 値が数値でない等の「そのセルでは成り立たない」理由は無視し、
                    # 式そのものが読めない場合だけ、あとで件数を知らせる
                    if ent["rule"].type == "expression" and ent["rule"].formula:
                        self.uneval.add(f"{str(ent['rule'].formula[0])[:40]} … {e}")
                    res = None
                except Exception:
                    res = None
                if res is None:
                    continue
                for k, v in res.items():
                    acc.setdefault(k, v)
                if ent["stop"]:
                    break
            if acc:
                self.map[(r, c)] = acc

    def get(self, r: int, c: int) -> dict | None:
        return self.map.get((r, c))

    # -- 集計（色スケール・データバー・上位/下位など） -----------------------
    def _agg(self, ent) -> dict:
        if ent["agg"] is not None:
            return ent["agg"]
        vals, texts = [], []
        budget = 200_000          # 最小/最大や重複判定のための走査量の上限
        for (r1, c1, r2, c2) in ent["full"]:
            for r in range(r1, r2 + 1):
                if budget <= 0:
                    break
                budget -= c2 - c1 + 1
                for c in range(c1, c2 + 1):
                    v = self.ws.cell(row=r, column=c).value
                    if v is None:
                        continue
                    texts.append(v.upper() if isinstance(v, str) else v)
                    if isinstance(v, bool) or isinstance(v, str):
                        continue
                    try:
                        vals.append(_num(v))
                    except EvalError:
                        pass
        vals.sort()
        counts: dict = {}
        for t in texts:
            counts[t] = counts.get(t, 0) + 1
        avg = sum(vals) / len(vals) if vals else 0.0
        var = sum((v - avg) ** 2 for v in vals) / len(vals) if vals else 0.0
        ent["agg"] = {"vals": vals, "min": vals[0] if vals else 0.0,
                      "max": vals[-1] if vals else 0.0, "avg": avg,
                      "sd": math.sqrt(var), "counts": counts}
        return ent["agg"]

    def _cfvo(self, cfvo, agg, ent, r, c, lo_hi=None):
        """<cfvo> → 実数のしきい値。"""
        t = getattr(cfvo, "type", None)
        raw = getattr(cfvo, "val", None)
        if t in ("min", "autoMin"):
            return agg["min"]
        if t in ("max", "autoMax"):
            return agg["max"]
        try:
            x = float(raw)
        except (TypeError, ValueError):
            if t == "formula" and raw:
                return _num(eval_formula(str(raw), EvalCtx(self.wb, self.ws, r, c,
                                                           ent["anchor"], self.today)))
            raise EvalError("cfvo が読めない")
        if t == "num":
            return x
        if t == "percent":
            lo, hi = (agg["min"], agg["max"]) if lo_hi is None else lo_hi
            return lo + (hi - lo) * x / 100.0
        if t == "percentile":
            vs = agg["vals"]
            if not vs:
                return 0.0
            pos = (len(vs) - 1) * x / 100.0
            i = int(math.floor(pos))
            j = min(i + 1, len(vs) - 1)
            return vs[i] + (vs[j] - vs[i]) * (pos - i)
        if t == "formula":
            return _num(eval_formula(str(raw), EvalCtx(self.wb, self.ws, r, c,
                                                       ent["anchor"], self.today)))
        return x

    # -- 差分書式 (dxf) -----------------------------------------------------
    def _dxf(self, rule) -> dict:
        out: dict = {}
        dxf = getattr(rule, "dxf", None)
        if dxf is None:
            return out
        f = getattr(dxf, "font", None)
        if f is not None:
            if f.b is not None:
                out["b"] = bool(f.b)
            if f.i is not None:
                out["i"] = bool(f.i)
            if f.strike is not None:
                out["strike"] = bool(f.strike)
            if f.u is not None:
                out["u"] = f.u if isinstance(f.u, str) else "single"
            col = self.R.resolve(f.color)
            if col:
                out["color"] = col
        fl = getattr(dxf, "fill", None)
        if fl is not None:
            # dxf の塗りは patternType を省略して bgColor に色を入れるのが Excel の書き方
            if getattr(fl, "patternType", None) == "solid":
                col = self.R.resolve(fl.fgColor) or self.R.resolve(fl.bgColor)
            else:
                col = self.R.resolve(fl.bgColor) or self.R.resolve(fl.fgColor)
            if col:
                out["bg"] = col
        bd = getattr(dxf, "border", None)
        if bd is not None:
            for side, key in (("top", "bt"), ("bottom", "bb"),
                              ("left", "bl"), ("right", "br")):
                css = border_css(getattr(bd, side, None), self.R)
                if css:
                    out[key] = css
        nf = getattr(dxf, "numFmt", None)
        if nf is not None and getattr(nf, "formatCode", None):
            out["numfmt"] = nf.formatCode
        return out

    # -- 1ルールをセルに当てる ----------------------------------------------
    def _apply(self, ent, r: int, c: int) -> dict | None:
        rule = ent["rule"]
        t = rule.type
        v = self.ws.cell(row=r, column=c).value
        blank = v is None or (isinstance(v, str) and v == "")

        if t == "colorScale" and rule.colorScale is not None:
            return self._color_scale(rule.colorScale, ent, r, c, v, blank)
        if t == "dataBar" and rule.dataBar is not None:
            return self._data_bar(rule.dataBar, ent, r, c, v, blank)
        if t == "iconSet" and rule.iconSet is not None:
            return self._icon_set(rule.iconSet, ent, r, c, v, blank)

        matched = self._match(rule, t, ent, r, c, v, blank)
        if not matched:
            return None
        return self._dxf(rule)

    def _match(self, rule, t, ent, r, c, v, blank) -> bool:
        if t == "expression":
            if not rule.formula:
                return False
            ctx = EvalCtx(self.wb, self.ws, r, c, ent["anchor"], self.today)
            return _truthy(eval_formula(str(rule.formula[0]), ctx))

        if t == "cellIs":
            if blank:
                return False
            ctx = EvalCtx(self.wb, self.ws, r, c, ent["anchor"], self.today)
            ops = [eval_formula(str(f), ctx) for f in (rule.formula or [])]
            if not ops:
                return False
            op = rule.operator
            a = ops[0]
            if op == "between":
                if len(ops) < 2:
                    return False
                lo, hi = (ops[0], ops[1])
                if _compare(lo, hi) > 0:
                    lo, hi = hi, lo
                return _compare(v, lo) >= 0 and _compare(v, hi) <= 0
            if op == "notBetween":
                if len(ops) < 2:
                    return False
                lo, hi = (ops[0], ops[1])
                if _compare(lo, hi) > 0:
                    lo, hi = hi, lo
                return not (_compare(v, lo) >= 0 and _compare(v, hi) <= 0)
            cmpv = _compare(v, a)
            return {"equal": cmpv == 0, "notEqual": cmpv != 0,
                    "greaterThan": cmpv > 0, "lessThan": cmpv < 0,
                    "greaterThanOrEqual": cmpv >= 0,
                    "lessThanOrEqual": cmpv <= 0}.get(op, False)

        if t in ("containsText", "notContainsText", "beginsWith", "endsWith"):
            s = _text(v).upper()
            needle = (rule.text or "").upper()
            if not needle:
                return False
            hit = {"containsText": needle in s,
                   "notContainsText": needle not in s,
                   "beginsWith": s.startswith(needle),
                   "endsWith": s.endswith(needle)}[t]
            return hit
        if t == "containsBlanks":
            return blank or (isinstance(v, str) and v.strip() == "")
        if t == "notContainsBlanks":
            return not (blank or (isinstance(v, str) and v.strip() == ""))
        if t == "containsErrors":
            return isinstance(v, str) and v in ERROR_TEXTS
        if t == "notContainsErrors":
            return not (isinstance(v, str) and v in ERROR_TEXTS)

        if t in ("duplicateValues", "uniqueValues"):
            if blank:
                return False
            agg = self._agg(ent)
            key = v.upper() if isinstance(v, str) else v
            n = agg["counts"].get(key, 0)
            return n > 1 if t == "duplicateValues" else n == 1

        if t == "top10":
            if blank or isinstance(v, str) or isinstance(v, bool):
                return False
            agg = self._agg(ent)
            vs = agg["vals"]
            if not vs:
                return False
            rank = int(rule.rank or 10)
            bottom = bool(rule.bottom)
            if rule.percent:
                k = max(int(len(vs) * rank / 100.0), 1)
            else:
                k = max(min(rank, len(vs)), 1)
            thr = vs[k - 1] if bottom else vs[len(vs) - k]
            x = _num(v)
            return x <= thr if bottom else x >= thr

        if t == "aboveAverage":
            if blank or isinstance(v, str) or isinstance(v, bool):
                return False
            agg = self._agg(ent)
            x, avg = _num(v), agg["avg"]
            sd = rule.stdDev
            above = rule.aboveAverage is None or bool(rule.aboveAverage)
            if sd:
                thr = avg + agg["sd"] * int(sd) * (1 if above else -1)
                return x > thr if above else x < thr
            if rule.equalAverage:
                return x >= avg if above else x <= avg
            return x > avg if above else x < avg

        if t == "timePeriod":
            if blank:
                return False
            try:
                d = _dt(v).date()
            except EvalError:
                return False
            td = self.today
            wd = td.weekday()                    # 月=0
            week0 = td - timedelta(days=(wd + 1) % 7)     # 今週の日曜
            p = rule.timePeriod
            if p == "today":
                return d == td
            if p == "yesterday":
                return d == td - timedelta(days=1)
            if p == "tomorrow":
                return d == td + timedelta(days=1)
            if p == "last7Days":
                return td - timedelta(days=6) <= d <= td
            if p == "thisWeek":
                return week0 <= d < week0 + timedelta(days=7)
            if p == "lastWeek":
                return week0 - timedelta(days=7) <= d < week0
            if p == "nextWeek":
                return week0 + timedelta(days=7) <= d < week0 + timedelta(days=14)
            if p == "thisMonth":
                return (d.year, d.month) == (td.year, td.month)
            if p == "lastMonth":
                m = td.month - 1 or 12
                y = td.year - (1 if td.month == 1 else 0)
                return (d.year, d.month) == (y, m)
            if p == "nextMonth":
                m = td.month + 1 if td.month < 12 else 1
                y = td.year + (1 if td.month == 12 else 0)
                return (d.year, d.month) == (y, m)
            return False

        self.unsupported.add(t or "?")
        return False

    # -- 色スケール / データバー / アイコンセット ---------------------------
    def _color_scale(self, cs, ent, r, c, v, blank):
        if blank or isinstance(v, str) or isinstance(v, bool):
            return None
        agg = self._agg(ent)
        x = _num(v)
        stops = []
        for i, cfvo in enumerate(cs.cfvo):
            thr = self._cfvo(cfvo, agg, ent, r, c)
            col = self.R.resolve(cs.color[i]) if i < len(cs.color) else None
            stops.append((thr, col or "#ffffff"))
        stops.sort(key=lambda s: s[0])
        if not stops:
            return None
        if x <= stops[0][0]:
            return {"bg": stops[0][1]}
        if x >= stops[-1][0]:
            return {"bg": stops[-1][1]}
        for i in range(len(stops) - 1):
            lo, hi = stops[i], stops[i + 1]
            if lo[0] <= x <= hi[0]:
                span = hi[0] - lo[0]
                t = 0.0 if span == 0 else (x - lo[0]) / span
                return {"bg": mix(hi[1], lo[1], t)}
        return None

    def _data_bar(self, db, ent, r, c, v, blank):
        if blank or isinstance(v, str) or isinstance(v, bool):
            return None
        agg = self._agg(ent)
        x = _num(v)
        cfvo = list(db.cfvo)
        lo = self._cfvo(cfvo[0], agg, ent, r, c) if cfvo else agg["min"]
        hi = self._cfvo(cfvo[-1], agg, ent, r, c) if len(cfvo) > 1 else agg["max"]
        lo, hi = min(lo, hi), max(lo, hi)
        lo = min(lo, 0.0) if agg["min"] < 0 else lo
        span = hi - lo
        pct = 0.0 if span == 0 else (x - lo) / span * 100.0
        # 省略時のスキーマ既定は 10/90 だが、いまの Excel は拡張(x14)側の設定で描くので
        # バーはセル幅いっぱいまで伸びる。明示されているときだけその値に従う。
        minlen = float(db.minLength if db.minLength is not None else 0)
        maxlen = float(db.maxLength if db.maxLength is not None else 100)
        pct = minlen + (maxlen - minlen) * max(min(pct, 100.0), 0.0) / 100.0
        col = self.R.resolve(db.color) or "#638ec6"
        out = {"bar": (mix(col, "#ffffff", 0.55), round(pct, 1),
                       bool(getattr(db, "direction", None) == "rightToLeft"))}
        if getattr(db, "showValue", True) is False:
            out["novalue"] = True
        return out

    def _icon_set(self, ics, ent, r, c, v, blank):
        if blank or isinstance(v, str) or isinstance(v, bool):
            return None
        agg = self._agg(ent)
        x = _num(v)
        cfvo = list(ics.cfvo)
        n = len(cfvo)
        idx = 0
        for i in range(1, n):
            try:
                thr = self._cfvo(cfvo[i], agg, ent, r, c)
            except EvalError:
                continue
            gte = getattr(cfvo[i], "gte", True)
            if (x >= thr) if (gte is None or gte) else (x > thr):
                idx = i
        setname = ics.iconSet or "3TrafficLights1"
        if getattr(ics, "reverse", False):
            idx = n - 1 - idx
        out = {"icon": (setname, idx)}
        if getattr(ics, "showValue", True) is False:
            out["novalue"] = True
        return out


_TAG_RE = re.compile(r"<[^>]*>")


def est_text_px(text: str, pt: float, bold: bool) -> float:
    """文字列の描画幅を『多めに』見積もる。はみ出しの可能性がある行だけを
    ブラウザ側で実測させるための足切りに使うので、過大評価する分には害がない。"""
    w = 0.0
    for ch in text:
        o = ord(ch)
        if o < 0x2E80 or 0xFF61 <= o <= 0xFF9F:
            w += 0.62          # 半角
        else:
            w += 1.02          # 全角
    if bold:
        w *= 1.06
    return w * pt * 4.0 / 3.0


def render_sheet(ws, resolver: ColorResolver, drawings: list[dict],
                 gridlines: bool, sheet_id: str, opts,
                 formulas: dict[tuple[int, int], str] | None = None,
                 bounds: tuple[int, int, int, int] | None = None,
                 page_breaks: set[int] | None = None,
                 title_rows: tuple[int, int] | None = None,
                 overrides: dict[tuple[int, int], str] | None = None,
                 cond: "CondFormat | None" = None) -> dict:
    full_row = min(ws.max_row or 1, opts.max_rows)
    full_col = min(ws.max_column or 1, opts.max_cols)
    grid = Grid(ws, full_col, full_row)
    min_row, min_col, max_row, max_col = bounds or (1, 1, full_row, full_col)
    max_row, max_col = min(max_row, full_row), min(max_col, full_col)
    styles = StyleSheet(resolver, gridlines)
    page_breaks = page_breaks or set()
    cfmap = cond.map if cond is not None else {}
    epoch = getattr(ws.parent, "epoch", None)     # 1904 年基準のブックがある

    # 結合セル（印刷範囲で切った場合は、はみ出した分を詰めて表示する）
    covered: dict[tuple[int, int], tuple[int, int]] = {}
    spans: dict[tuple[int, int], tuple[int, int]] = {}
    src: dict[tuple[int, int], tuple[int, int]] = {}
    for rng in ws.merged_cells.ranges:
        r1, c1, r2, c2 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        if r1 > max_row or c1 > max_col or r2 < min_row or c2 < min_col:
            continue
        ar, ac = max(r1, min_row), max(c1, min_col)
        er, ec = min(r2, max_row), min(c2, max_col)
        spans[(ar, ac)] = (er - ar + 1, ec - ac + 1)
        if (ar, ac) != (r1, c1):
            src[(ar, ac)] = (r1, c1)
        for r in range(ar, er + 1):
            for c in range(ac, ec + 1):
                if (r, c) != (ar, ac):
                    covered[(r, c)] = (ar, ac)

    # 1パス目: 表示文字列を確定し、行ごとの「文字が入っているセル」を把握する
    #（Excel は隣が空セルのときだけ文字をはみ出させるため、その判定に使う）
    content: dict[tuple[int, int], str] = {}
    plain: dict[tuple[int, int], str] = {}
    occupied: dict[int, set[int]] = {}
    formulas = formulas or {}
    overrides = overrides or {}
    missing_cache = 0
    for r in range(min_row, max_row + 1):
        occ = set()
        for c in range(min_col, max_col + 1):
            if (r, c) in covered:
                occ.add(c)
                continue
            sr, sc = src.get((r, c), (r, c))
            cell = ws.cell(row=sr, column=sc)

            def put(html_text: str, raw: str):
                content[(r, c)] = html_text
                plain[(r, c)] = raw
                occ.add(c)
                rs_, cs_ = spans.get((r, c), (1, 1))
                for cc in range(c, min(c + cs_, max_col + 1)):
                    occ.add(cc)

            ov = overrides.get((sr, sc))
            if ov is not None:      # Excel なら再計算されて出るはずの文字列
                put(esc(ov), ov)
                continue
            if cell.value is None:
                # 数式なのに計算結果が保存されていないセル
                fml = formulas.get((sr, sc))
                if not fml:
                    continue
                missing_cache += 1
                if not opts.show_formula:
                    continue
                put(esc(tidy_formula(fml)), tidy_formula(fml))
                continue
            hval = rich_text_html(cell.value)
            if hval is None:
                cfx = cfmap.get((sr, sc))
                nfmt = (cfx or {}).get("numfmt") or cell.number_format
                text = format_cell_value(cell.value, nfmt, epoch)
                hval = esc(text) if text != "" else ""
                raw = text
            else:
                raw = _TAG_RE.sub("", hval)
            if hval == "":
                continue
            put(hval, raw)
        occupied[r] = occ

    def has_border(r: int, c: int, side: str) -> bool:
        if not (min_row <= r <= max_row and min_col <= c <= max_col):
            return False
        b = ws.cell(row=r, column=c).border
        s = getattr(b, side, None) if b is not None else None
        return bool(s is not None and s.style)

    def grid_edges(r: int, c: int, rs: int, cs: int) -> tuple[bool, bool]:
        """右辺・下辺に目盛線を引いてよいか。隣に罫線があるなら引かない。"""
        if not gridlines:
            return True, True
        rb = r + rs
        while rb in grid.hidden_rows:            # 非表示行は出力しないので、その先を見る
            rb += 1
        return (not any(has_border(rr, c + cs, "left") for rr in range(r, r + rs)),
                not any(has_border(rb, cc, "top") for cc in range(c, c + cs)))

    def free_width(r: int, c_from: int, step: int) -> tuple[float, bool]:
        """隣接する空セルの合計幅と、行端まで空だったかを返す。"""
        w = 0.0
        occ = occupied.get(r, ())
        c = c_from
        while min_col <= c <= max_col:
            if c in occ:
                return w, True
            w += grid.col_px[c]
            c += step
        return w, False

    # 印刷範囲で切った場合は、その左上を原点にする
    x0, y0 = grid.x_at(min_col - 1), grid.y_at(min_row - 1)
    table_w = sum(grid.col_px[min_col:max_col + 1])
    table_h = sum(grid.row_px[min_row:max_row + 1])

    out = []
    cw, ch = table_w, table_h
    shapes = []
    for s in drawings:                     # 図形が用紙外に出る場合は台紙を広げる
        x, y, w, h = s["box"]
        x, y = x - x0, y - y0
        if x + w < 0 or y + h < 0:
            continue
        s = dict(s, box=(x, y, w, h))
        shapes.append(s)
        cw = max(cw, x + w + 2)
        ch = max(ch, y + h + 2)
    out.append(f'<div class="canvas" style="width:{cw:.0f}px;height:{ch:.0f}px">')
    out.append(f'<table style="width:{table_w:.0f}px"><colgroup>')
    for c in range(min_col, max_col + 1):
        out.append(f'<col style="width:{grid.col_px[c]}px">')
    out.append("</colgroup>")

    # 同じ式が縦や横にコピーされている範囲は、先頭のセルにだけ吹き出しを出す。
    # 方眼紙の表では同じ式が何十行も続くので、全部出すと吹き出しで埋まってしまう。
    # 隠すのは CSS 側なので、画面下部のトグルで残りも出せる。
    fb_dup: set[tuple[int, int]] = set()
    fb_rep: dict[tuple[int, int], int] = {}
    if opts.formula_balloons and formulas:
        shape: dict[tuple[int, int], str] = {}
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) in covered:
                    continue
                sr, sc = src.get((r, c), (r, c))
                fml = formulas.get((sr, sc))
                if fml:
                    shape[(r, c)] = formula_shape(fml, sr, sc)
        head_of: dict[tuple[int, int], tuple[int, int]] = {}
        up: dict[int, tuple[int, int]] = {}      # 列 → その列で直前に見た数式セル
        for r in range(min_row, max_row + 1):
            left: tuple[int, int] | None = None
            for c in range(min_col, max_col + 1):
                s = shape.get((r, c))
                if s is None:
                    continue
                head = None
                for prev in (up.get(c), left):   # 上と左のどちらかの続きなら繰り返し
                    if prev is not None and shape.get(prev) == s:
                        head = head_of[prev]
                        break
                if head is None:
                    head_of[(r, c)] = (r, c)
                    fb_rep[(r, c)] = 1
                else:
                    head_of[(r, c)] = head
                    fb_dup.add((r, c))
                    fb_rep[head] += 1
                up[c] = left = (r, c)

    n_balloon = n_hidden = 0

    def balloon(r: int, c: int, sr: int, sc: int) -> str:
        """数式の吹き出し（--formula-balloons）。数式セルの下に常時出す。"""
        nonlocal n_balloon, n_hidden
        fml = formulas.get((sr, sc))
        if not fml:
            return ""
        n_balloon += 1
        body, cols, rows = format_formula(fml)
        rep = fb_rep.get((r, c), 1)
        if rep > 1:              # 同じ式が続く範囲の先頭。何セル分かを添える
            body += f'<b class="c"> ×{rep}</b>'
        kls = "fb"
        w = min(cols, FB_WRAP) * FB_CH_PX + 14
        x = grid.x_at(c - 1) - x0
        if x + w > cw and x > w:        # 台紙の右端をはみ出すので右寄せにする
            kls += " fbr"
        if rows > FB_LINES:
            kls += " fbc"              # 畳んでおき、ホバー／クリックで全文を出す
        if (r, c) in fb_dup:
            kls += " fbd"              # 同じ式の繰り返し。既定では隠す
            n_hidden += 1
        # 重なったときに上・左のセルの吹き出しが手前に来るようにする
        z = 20 + (max_row - r) * 2048 + (max_col - c)
        return f'<div class="{kls}" style="--fbz:{z}">{body}</div>'

    thead_end = 0
    if title_rows and title_rows[0] <= min_row:
        thead_end = min(title_rows[1], max_row)
    out.append("<thead>" if thead_end >= min_row else "<tbody>")

    for r in range(min_row, max_row + 1):
        if r in grid.hidden_rows:
            continue
        if thead_end and r == thead_end + 1:
            out.append("</thead><tbody>")
        brk = ' class="pb"' if r in page_breaks and r > min_row else ""
        out.append(f'<tr{brk} style="height:{grid.row_px[r]:.0f}px">')
        for c in range(min_col, max_col + 1):
            if (r, c) in covered:
                continue
            if grid.col_px[c] == 0 and (r, c) not in spans:
                out.append('<td class="hidden"></td>')
                continue
            sr, sc = src.get((r, c), (r, c))
            cell = ws.cell(row=sr, column=sc)
            rs, cs = spans.get((r, c), (1, 1))
            cfx = cfmap.get((sr, sc))
            cls = styles.class_for(cell, *grid_edges(r, c, rs, cs), cfx)
            fb = balloon(r, c, sr, sc) if opts.formula_balloons else ""
            attr = f' class="{cls} fx"' if fb else f' class="{cls}"'
            if rs > 1:
                attr += f' rowspan="{rs}"'
            if cs > 1:
                attr += f' colspan="{cs}"'
            body = content.get((r, c))
            icon = ""
            if cfx and cfx.get("icon"):
                icon = icon_svg(*cfx["icon"])
            if not body and not icon:
                out.append(f"<td{attr}>{fb}</td>")
                continue
            body = body or ""

            # 見切れの再現。Excel と同じく、横は「隣に文字があるところ」まで、
            # 縦は行の高さまでで切る。切った文字も DOM には残すのでコピー・検索できる。
            al = cell.alignment
            wrap = bool(al and al.wrapText)
            halign = (al.horizontal if al else None) or ""
            own = sum(grid.col_px[c:c + cs])
            limit, kls = "", ""
            if wrap:
                own_h = sum(grid.row_px[r:r + rs])
                limit = f' style="max-height:{own_h:.0f}px"'
                kls = " cw"
            elif not (al and al.textRotation):
                if halign in ("", "left", "general", "justify", "distributed", "fill"):
                    avail, blocked = free_width(r, c + cs, 1)
                    if blocked:
                        limit = f' style="max-width:{own + avail:.0f}px"'
                        kls = " c"
                elif halign in ("center", "centerContinuous"):
                    la, lb = free_width(r, c - 1, -1)
                    ra, rb = free_width(r, c + cs, 1)
                    if lb or rb:
                        limit = f' style="max-width:{own + 2 * min(la, ra):.0f}px"'
                        kls = " c"
                elif halign == "right":
                    avail, blocked = free_width(r, c - 1, -1)
                    if blocked:
                        limit = f' style="max-width:{own + avail:.0f}px"'
                        kls = " cr"
                # はみ出す見込みがあるセルに印を付ける。Excel では、はみ出した文字は
                # 通り道の目盛線を消すので、実際にどこまで届くかの実測は
                # ブラウザ側（mark()）に任せる。ov=右へ / ovl=左へ / ovc=両側へ
                if gridlines:
                    fpt = float(cell.font.sz or 11) if cell.font else 11.0
                    if est_text_px(plain.get((r, c), ""), fpt,
                                   bool(cell.font and cell.font.b)) > own - 2 * CELL_PAD:
                        kls += {"right": " ovl", "center": " ovc",
                                "centerContinuous": " ovc"}.get(halign, " ov")
            if cfx and cfx.get("novalue"):
                kls += " nv"
            if kls:
                kls = f' class="{kls.strip()}"'

            link = cell.hyperlink
            if link is not None and getattr(link, "target", None):
                inner = icon + f'<a href="{esc(link.target)}">{body}</a>'
            else:
                inner = icon + body
            tips = []
            if cell.comment is not None and cell.comment.text:
                tips.append(cell.comment.text)
            if opts.formula_tips:
                fml = formulas.get((sr, sc))
                if fml:
                    tips.append(tidy_formula(fml))
            tip = f' title="{esc(chr(10).join(tips))}"' if tips else ""
            out.append(f"<td{attr}{tip}>{fb}<i{kls}{limit}>{inner}</i></td>")
        out.append("</tr>")
    out.append("</tbody></table>")

    if shapes:
        out.append('<div class="draw">')
        for s in shapes:
            out.append(render_shape(s))
        out.append("</div>")
    out.append("</div>")
    return {"html": "".join(out), "css": styles.css(), "missing": missing_cache,
            "width": cw, "height": ch, "rows": max_row - min_row + 1,
            "cols": max_col - min_col + 1, "balloons": n_balloon, "fbdup": n_hidden}


# 矢尻の大きさ（OOXML の sm / med / lg）。線幅の倍数で効く
ARROW_SCALE = {"sm": 0.7, "med": 1.0, "lg": 1.4}


def arrow_marker(uid: str, end, color: str, at_end: bool) -> str:
    """<a:headEnd>/<a:tailEnd> → SVG の <marker>。返り値が空なら矢尻なし。"""
    if not end:
        return ""
    kind, wv, lv = end if isinstance(end, tuple) else (end, None, None)
    if not kind or kind == "none":
        return ""
    fw = 5.0 * ARROW_SCALE.get(wv or "med", 1.0)      # 幅（線幅の倍数）
    fl = 5.0 * ARROW_SCALE.get(lv or "med", 1.0)      # 長さ
    cy = fw / 2
    # marker の座標系は「線の進行方向 = +x」。終端は先端を右に、始端は左に置く
    if kind == "triangle":
        d, fillmode = f"M0,0 L{fl},{cy} L0,{fw} Z", "fill"
    elif kind == "stealth":
        d, fillmode = f"M0,0 L{fl},{cy} L0,{fw} L{fl*0.35:.2f},{cy} Z", "fill"
    elif kind == "arrow":
        d, fillmode = f"M0,0 L{fl},{cy} L0,{fw}", "stroke"
    elif kind == "diamond":
        d, fillmode = (f"M0,{cy} L{fl/2:.2f},0 L{fl},{cy} L{fl/2:.2f},{fw} Z", "fill")
    elif kind == "oval":
        d, fillmode = "", "oval"
    else:
        d, fillmode = f"M0,0 L{fl},{cy} L0,{fw} Z", "fill"

    if fillmode == "oval":
        shape = (f'<circle cx="{fl/2:.2f}" cy="{cy:.2f}" r="{min(fl,fw)/2:.2f}" '
                 f'fill="{color}"/>')
    elif fillmode == "stroke":
        shape = f'<path d="{d}" fill="none" stroke="{color}" stroke-width="1"/>'
    else:
        shape = f'<path d="{d}" fill="{color}"/>'
    # 先端がちょうど線の端点に来るよう refX を合わせる
    ref_x = fl if at_end else 0.0
    tr = "" if at_end else f' transform="rotate(180 {fl/2:.2f} {cy:.2f})"'
    if tr:
        ref_x = 0.0
    return (f'<marker id="{uid}" markerWidth="{fl:.2f}" markerHeight="{fw:.2f}" '
            f'refX="{ref_x:.2f}" refY="{cy:.2f}" orient="auto">'
            f'<g{tr}>{shape}</g></marker>')


def render_shape(s: dict) -> str:
    x, y, w, h = s["box"]
    w = max(w, 1.0)
    h = max(h, 1.0)
    base = f"left:{x:.1f}px;top:{y:.1f}px;width:{w:.1f}px;height:{h:.1f}px"
    rot = s.get("rot") or 0
    if rot:
        base += f";transform:rotate({rot:.2f}deg)"
    if s["kind"] == "img":
        alt = esc(s.get("alt") or "")
        return (f'<img class="dobj" style="{base}" src="{s["src"]}" alt="{alt}">')

    prst = s.get("prst", "rect")
    path = shape_path(prst, w, h, s.get("adj"))
    fill, line, lw = s.get("fill"), s.get("line"), s.get("line_w") or 0
    dash = s.get("dash")
    body = []

    is_line = is_line_shape(prst)
    if is_line:
        fill = None                 # 開いたパスなので塗ると塊になる
    if path is not None:
        stroke = line or ("#000000" if is_line else "none")
        sw = lw if lw else (1 if is_line else 0)
        dasharray = ""
        if dash in ("dash", "sysDash", "lgDash"):
            dasharray = f' stroke-dasharray="{max(sw*3,4)},{max(sw*2,3)}"'
        elif dash in ("dot", "sysDot"):
            dasharray = f' stroke-dasharray="{max(sw,1)},{max(sw*2,2)}"'
        elif dash in ("dashDot", "sysDashDot", "lgDashDot"):
            dasharray = f' stroke-dasharray="{max(sw*3,4)},{max(sw*2,3)},{max(sw,1)},{max(sw*2,3)}"'
        # 反転
        tf = ""
        if s.get("flip_h") or s.get("flip_v"):
            sx, sy = (-1 if s.get("flip_h") else 1), (-1 if s.get("flip_v") else 1)
            cx, cy = w / 2, h / 2
            tf = f' transform="translate({cx if sx<0 else 0},{cy if sy<0 else 0}) scale({sx},{sy}) translate({-cx if sx<0 else 0},{-cy if sy<0 else 0})"'
        marker = ""
        defs = ""
        uid = f"m{abs(hash((x, y, w, h, prst))) % 100000}"
        d = arrow_marker(f"{uid}t", s.get("tail"), line or "#000", True)
        if d:
            defs += d
            marker += f' marker-end="url(#{uid}t)"'
        d = arrow_marker(f"{uid}h", s.get("head"), line or "#000", False)
        if d:
            defs += d
            marker += f' marker-start="url(#{uid}h)"'
        body.append(
            f'<svg class="dsvg" width="{w:.1f}" height="{h:.1f}" viewBox="0 0 {w:.1f} {h:.1f}" '
            f'overflow="visible">' +
            (f"<defs>{defs}</defs>" if defs else "") +
            f'<path d="{path}" fill="{fill or "none"}" stroke="{stroke}" '
            f'stroke-width="{sw}"{dasharray}{marker}{tf}/></svg>')
        shell_style = base
    else:
        # 矩形系は CSS で
        st = [base]
        if fill:
            st.append(f"background:{fill}")
        if line and lw:
            st.append(f"border:{lw:.1f}px {'dashed' if dash and 'dash' in dash else 'solid'} {line}")
        if prst in ("roundRect", "round1Rect", "round2SameRect", "snip1Rect"):
            st.append(f"border-radius:{min(w, h) * 0.16:.1f}px")
        elif prst in ("ellipse", "flowChartConnector"):
            st.append("border-radius:50%")
        elif prst == "can":
            st.append("border-radius:50% / 12%")
        shell_style = ";".join(st)

    txt = s.get("text")
    text_html = ""
    if txt:
        anchor = {"t": "flex-start", "ctr": "center", "b": "flex-end"}.get(txt.get("anchor") or "ctr",
                                                                          "center")
        vert = txt.get("vert")
        vcss = ""
        if vert in ("vert", "eaVert"):
            vcss = "writing-mode:vertical-rl;"
        elif vert == "vert270":
            vcss = "writing-mode:vertical-rl;transform:rotate(180deg);"
        paras = []
        for p in txt["paras"]:
            algn = {"l": "left", "ctr": "center", "r": "right", "just": "justify"}.get(
                p.get("algn") or "ctr", "center")
            runs = []
            for run in p["runs"]:
                if run["text"] == "\n":
                    runs.append("<br>")
                    continue
                css = []
                if run.get("size"):
                    css.append(f'font-size:{run["size"]}pt')
                if run.get("b"):
                    css.append("font-weight:700")
                if run.get("i"):
                    css.append("font-style:italic")
                if run.get("u"):
                    css.append("text-decoration:underline")
                if run.get("color"):
                    css.append(f'color:{run["color"]}')
                if run.get("font"):
                    css.append(f'font-family:"{run["font"]}",{FONT_FALLBACK}')
                t = esc(run["text"])
                runs.append(f'<span style="{";".join(css)}">{t}</span>' if css else t)
            paras.append(f'<p style="text-align:{algn}">{"".join(runs)}</p>')
        text_html = (f'<div class="dtx" style="align-items:{anchor};{vcss}">'
                     f'{"".join(paras)}</div>')

    return f'<div class="dobj" style="{shell_style}">{"".join(body)}{text_html}</div>'


# ---------------------------------------------------------------------------
# 印刷設定（用紙・余白・改ページ・印刷範囲）
# ---------------------------------------------------------------------------

# Excel の用紙コード → (幅, 高さ) インチ
PAPER = {
    1: (8.5, 11), 2: (8.5, 11), 5: (8.5, 14), 7: (7.25, 10.5),
    8: (11.69, 16.54),    # A3
    9: (8.27, 11.69),     # A4
    11: (5.83, 8.27),     # A5
    12: (10.12, 14.33),   # B4 (JIS)
    13: (7.17, 10.12),    # B5 (JIS)
    43: (10.12, 14.33), 44: (7.17, 10.12),
}


def print_area_bounds(ws, opts) -> tuple[int, int, int, int] | None:
    """印刷範囲を (min_row, min_col, max_row, max_col) にする。複数指定は外接矩形。"""
    area = ws.print_area
    if not area:
        return None
    parts = area if isinstance(area, (list, tuple)) else [area]
    rows, cols = [], []
    for p in parts:
        for token in str(p).split(","):
            token = token.split("!")[-1].replace("$", "").strip()
            if not token:
                continue
            try:
                from openpyxl.utils import range_boundaries
                c1, r1, c2, r2 = range_boundaries(token)
            except Exception:
                continue
            if None in (c1, r1, c2, r2):
                continue
            rows += [r1, r2]
            cols += [c1, c2]
    if not rows:
        return None
    return (min(rows), min(cols), min(max(rows), opts.max_rows),
            min(max(cols), opts.max_cols))


def page_settings(ws, sheet_width_px: float) -> dict:
    """@page 用の用紙・余白・拡大率を組み立てる。"""
    ps, pm = ws.page_setup, ws.page_margins
    w_in, h_in = PAPER.get(ps.paperSize or 9, PAPER[9])
    landscape = (ps.orientation or "portrait") == "landscape"
    if landscape:
        w_in, h_in = h_in, w_in
    ml = pm.left if pm.left is not None else 0.7
    mr = pm.right if pm.right is not None else 0.7
    mt = pm.top if pm.top is not None else 0.75
    mb = pm.bottom if pm.bottom is not None else 0.75

    zoom = 1.0
    fit = getattr(ws.sheet_properties.pageSetUpPr, "fitToPage", False) \
        if ws.sheet_properties.pageSetUpPr is not None else False
    if fit:
        fw = ps.fitToWidth if ps.fitToWidth is not None else 1
        if fw:
            printable = (w_in - ml - mr) * 96.0
            if sheet_width_px > printable > 0:
                zoom = printable / sheet_width_px
    elif ps.scale:
        zoom = ps.scale / 100.0
    return {"w": w_in * 25.4, "h": h_in * 25.4, "m": (mt * 25.4, mr * 25.4, mb * 25.4, ml * 25.4),
            "zoom": zoom}


def title_row_range(ws) -> tuple[int, int] | None:
    t = ws.print_title_rows
    if not t:
        return None
    m = re.match(r"\$?(\d+):\$?(\d+)", str(t).split("!")[-1].replace("$", ""))
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


# ---------------------------------------------------------------------------
# 出力全体
# ---------------------------------------------------------------------------

PAGE_CSS = """
:root{color-scheme:light;--gl:%(grid)s}
body.nogl{--gl:transparent}
*{box-sizing:border-box}
body{margin:0;background:#f3f3f3;font-family:%(ff)s;color:#000}
/* Excel と同じくシートタブは常に画面の下端に貼り付ける（DOM 上も本文のあと）。
   本文がタブに隠れないよう、body の下余白はタブの高さに合わせて JS で調整する */
.tabs{position:fixed;left:0;right:0;bottom:0;z-index:2000000000;
  display:flex;gap:2px;padding:0 8px 6px;
  background:#f8f8f8;border-top:1px solid #d0d0d0;flex-wrap:wrap}
.tabs button{font:inherit;font-size:12px;padding:5px 14px;border:1px solid #cfcfcf;
  border-top:none;border-radius:0 0 5px 5px;background:#ececec;cursor:pointer;color:#333}
.tabs button.on{background:#fff;color:#107c41;font-weight:700;box-shadow:inset 0 -2px 0 #107c41}
.tabs .tg{margin-left:auto;align-self:center;font-size:12px;color:#444;padding:5px 6px 0;
  display:flex;gap:5px;align-items:center;cursor:pointer;user-select:none}
.tabs .tg~.tg{margin-left:14px}   /* 2つ目以降は右端に並べる */
.wrap{padding:16px 48px 72px;overflow:auto}
.sheet{display:none}
.sheet.on{display:block}
.canvas{position:relative;background:#fff;box-shadow:0 1px 4px rgba(0,0,0,.18)}
table{border-collapse:collapse;table-layout:fixed;border-spacing:0;width:auto}
td{padding:0;position:relative;overflow:visible;vertical-align:bottom;
  font-size:11pt;line-height:1.18}
td.hidden{border:none!important;font-size:0}
td>i{font-style:normal;display:inline-block;z-index:1;padding:0 %(pad)spx;
  overflow:hidden;pointer-events:auto}
td>i.cr{display:flex;justify-content:flex-end}
/* Excel と同じく、塗りやはみ出した文字が通った目盛線は消える。
   --glr/--glb は目盛線の色だけを差し替える穴なので、本物の罫線には影響しない */
td.ngr{--glr:transparent}
svg.cfi{vertical-align:-0.14em;margin-right:2px;flex:none}
td>i.nv{color:transparent}          /* データバー/アイコンで「値を表示しない」設定 */
/* 見切れているセル: クリックで全文表示を固定、もう一度クリックで元に戻す。
   切れている文字も DOM 上には残っているので、開かなくても選択コピー・検索はできる。 */
td>i.on{cursor:zoom-in}
td>i.on.open{cursor:zoom-out}
body.mark td>i.on:not(.open){box-shadow:inset -3px 0 0 rgba(198,74,60,.5)}
body.mark td>i.cr.on:not(.open){box-shadow:inset 3px 0 0 rgba(198,74,60,.5)}
body.mark td>i.cw.on:not(.open){box-shadow:inset 0 -3px 0 rgba(198,74,60,.5)}
td>i.on.open{overflow:visible;max-width:none!important;max-height:none!important;
  z-index:30;background:#fffdf0;box-shadow:0 0 0 1px #d9b95c,0 2px 8px rgba(0,0,0,.2)}
.measuring td>i{max-width:none!important;max-height:none!important}
/* 折り返しセルは絶対配置に切り替えて開く（行の高さを押し広げないため） */
td>i.cw.on.open{position:absolute;top:0;left:0;right:0;width:auto}
td a{color:#0563c1}
/* 数式の吹き出し（--formula-balloons）。数式チェック用に数式セルの下へ常時出す。
   小さめの等幅で、長い式は Python 側で括弧と引数区切りの位置で折ってある。
   それでも入らない行はブラウザ側で折り返す（overflow-wrap）。 */
body.nofb .fb{display:none}
body:not(.fbdup) .fb.fbd{display:none}     /* 上や左と同じ式の繰り返しは隠す */
.fb{position:absolute;left:0;top:100%%;margin-top:5px;
  /* 重なったときに上・左のセルの吹き出しが手前に来るよう、行と列から z を決める */
  z-index:var(--fbz,20);
  width:max-content;max-width:calc(%(fbw)sch + 12px);padding:2px 5px;
  font:400 10px/1.45 ui-monospace,SFMono-Regular,Consolas,"Courier New",monospace;
  white-space:pre-wrap;overflow-wrap:anywhere;text-align:left;
  color:#333;background:#fffbe6;border:1px solid #d9b95c;border-radius:3px;
  box-shadow:0 1px 3px rgba(0,0,0,.16);cursor:zoom-in}
.fb.fbr{left:auto;right:0}                          /* 台紙の右端をはみ出す位置 */
.fb::before{content:"";position:absolute;top:-4px;left:9px;width:6px;height:6px;
  background:#fffbe6;border-left:1px solid #d9b95c;border-top:1px solid #d9b95c;
  transform:rotate(45deg)}
.fb.fbr::before{left:auto;right:9px}
/* 重なりを避けて下へずらした吹き出しを、元のセルにつなぐ引き出し線 */
.fb .fbl{position:absolute;left:11px;bottom:100%%;width:1px;background:#d9b95c;
  pointer-events:none}
.fb.fbr .fbl{left:auto;right:11px}
.fb.fbc{max-height:%(fbh)sem;overflow:hidden}       /* 長すぎる式は畳んでおく */
.fb.fbc::after{content:"";position:absolute;left:0;right:0;bottom:0;height:1.45em;
  background:linear-gradient(rgba(255,251,230,0),#fffbe6)}
.fb:hover,.fb.open{max-height:none;z-index:1999999999;cursor:zoom-out;
  box-shadow:0 2px 10px rgba(0,0,0,.3)}
.fb:hover::after,.fb.open::after{display:none}
.fb b{font-weight:inherit}
.fb .t{color:#a31515}      /* 文字列 */
.fb .n{color:#098658}      /* 数値・配列定数 */
.fb .f{color:#795e26;font-weight:600}   /* 関数名 */
.fb .r{color:#0070c1}      /* セル参照・構造参照 */
.fb .h{color:#af00db}      /* シート名 */
.fb .e{color:#c64a3c;font-weight:700}   /* エラー値 */
.fb .p0{color:#0e8a9e}     /* 括弧は深さで色を変えて対応を追いやすくする */
.fb .p1{color:#c26a00}
.fb .p2{color:#8b5cf6}
.fb .c{color:#8a8a8a}      /* 「×12」= 同じ式が続くセル数 */
body.fxm td.fx::after{content:"";position:absolute;top:0;right:0;z-index:6;
  border:3px solid transparent;border-top-color:#c64a3c;border-right-color:#c64a3c}
.draw{position:absolute;inset:0;pointer-events:none;z-index:5}
.dobj{position:absolute;pointer-events:auto}
img.dobj{object-fit:fill}
.dsvg{position:absolute;left:0;top:0;overflow:visible}
.dtx{position:absolute;inset:0;display:flex;flex-direction:column;justify-content:center;
  padding:3px 5px;font-size:10.5pt;line-height:1.25}
.dtx p{margin:0}
@media print{
  :root{--gl:transparent}          /* Excel と同じく目盛線は印刷しない */
  body{background:#fff;padding-bottom:0!important}
  .tabs{display:none}
  .wrap{padding:0;overflow:visible}
  .sheet{display:block!important;break-after:page}
  .sheet:last-child{break-after:auto}
  .canvas{box-shadow:none}
  tr.pb{break-before:page}          /* Excel の改ページ位置 */
  td>i.on{box-shadow:none}
  .fb{box-shadow:none}
}
"""


def build_html(title: str, sheets: list[dict]) -> str:
    css = [PAGE_CSS % {"ff": FONT_FALLBACK, "pad": CELL_PAD, "grid": GRID_COLOR,
                       "fbw": FB_WRAP, "fbh": f"{FB_LINES * 1.45:.2f}"}]
    for i, s in enumerate(sheets):
        # シートごとにクラスで名前空間を切る（id で切ると詳細度が高すぎて
        # 見切れ表示などの上書きが効かなくなる）
        scoped = []
        for line in s["css"].split("\n"):
            if line:
                scoped.append(f".s-{i} " + line)
        css.append("\n".join(scoped))
        pg = s.get("page")
        if pg:
            mt, mr, mb, ml = pg["m"]
            css.append(f"@page p{i}{{size:{pg['w']:.0f}mm {pg['h']:.0f}mm;"
                       f"margin:{mt:.0f}mm {mr:.0f}mm {mb:.0f}mm {ml:.0f}mm}}\n"
                       f".s-{i}{{page:p{i}}}")
            if abs(pg["zoom"] - 1.0) > 0.01:
                css.append(f"@media print{{.s-{i} .canvas{{zoom:{pg['zoom']:.4f}}}}}")
        if s.get("print_gridlines"):     # Excel 側で「枠線を印刷する」が有効なシート
            css.append(f"@media print{{.s-{i}{{--gl:{GRID_COLOR}}}}}")
    tabs = "".join(
        f'<button data-i="{i}" class="{"on" if i == 0 else ""}">{esc(s["name"])}</button>'
        for i, s in enumerate(sheets))
    bodies = "".join(
        f'<div class="sheet s-{i}{" on" if i == 0 else ""}" id="sh{i}">{s["html"]}</div>'
        for i, s in enumerate(sheets))
    script = """
function mark(sh){
  if(!sh||sh.dataset.marked||!sh.classList.contains('on'))return;
  sh.dataset.marked='1';
  // 右揃えの見切れは scrollWidth に出ないので、制限を外した幅と比べて判定する
  var cand=[].slice.call(sh.querySelectorAll('td>i.c,td>i.cr,td>i.ov,td>i.ovl,td>i.ovc'));
  var lim=cand.map(function(el){return parseFloat(el.style.maxWidth)||0});
  sh.classList.add('measuring');
  var box=cand.map(function(el){return el.getBoundingClientRect()});
  sh.classList.remove('measuring');
  cand.forEach(function(el,i){if(lim[i]&&box[i].width>lim[i]+2)el.classList.add('on')});
  sh.querySelectorAll('td>i.cw').forEach(function(el){
    if(el.scrollHeight>el.clientHeight+2||el.scrollWidth>el.clientWidth+2)el.classList.add('on');
  });
  // Excel でははみ出した文字が通り道の目盛線を消す。文字がどこまで届くかは
  // フォント次第なので、ここで実測してから消す。読み取りを全部済ませてから
  // まとめて書き込み、レイアウトの往復を避ける。
  var kill=[];
  cand.forEach(function(el,i){
    var cl=el.classList;
    var toR=cl.contains('ov')||cl.contains('ovc'), toL=cl.contains('ovl')||cl.contains('ovc');
    if(!toR&&!toL)return;
    var td=el.parentNode, cb=td.getBoundingClientRect(), lo=box[i].left+2, hi=box[i].right-2;
    if(lim[i]){                       // 見切れているセルは切れた先までしか消さない
      if(cl.contains('ovc')){var mid=(cb.left+cb.right)/2;
        lo=Math.max(lo,mid-lim[i]/2); hi=Math.min(hi,mid+lim[i]/2);}
      else if(cl.contains('ovl')) lo=Math.max(lo,cb.right-lim[i]);
      else hi=Math.min(hi,cb.left+lim[i]);
    }
    var t;
    if(toR){t=td; while(t&&t.getBoundingClientRect().right<hi){kill.push(t);t=t.nextElementSibling;}}
    if(toL){t=td.previousElementSibling;
      while(t&&t.getBoundingClientRect().right>lo){kill.push(t);t=t.previousElementSibling;}}
  });
  kill.forEach(function(t){t.classList.add('ngr')});
}
// 数式の吹き出しが完全に隠れないように、重なったものを下へずらす。手前に来るもの
// （上の行・左の列）は元の位置に残し、後ろに回るものだけを動かして上端を覗かせる。
// ずらした分は元のセルまで引き出し線でつなぐ。
function place(sh){
  if(!sh)return;
  var band=15;                       // 少なくともこの高さは見えるようにする
  var fbs=[].slice.call(sh.querySelectorAll('.fb'));
  fbs.forEach(function(el){
    el.style.marginTop='';
    var l=el.querySelector('.fbl'); if(l)l.remove();
  });
  fbs=fbs.filter(function(el){return el.offsetParent!==null});   // 隠れているものは除く
  if(!fbs.length||fbs.length>3000)return;    // 多すぎるときは計算が重いので諦める
  fbs.sort(function(a,b){
    return (parseInt(b.style.getPropertyValue('--fbz'),10)||0)
         - (parseInt(a.style.getPropertyValue('--fbz'),10)||0);
  });
  // 動かす前にまとめて測る（絶対配置なので、ずらしても他の位置は変わらない）
  var box=fbs.map(function(el){return el.getBoundingClientRect()});
  var placed=[],low=0;
  fbs.forEach(function(el,i){
    var r=box[i],t=r.top,g=0;
    while(g++<60){                   // 重なった相手の下まで飛ばす
      var hit=null;
      for(var j=placed.length-1;j>=0;j--){   // 近いものから見る
        var p=placed[j];
        if(p.b>t&&p.t<t+band&&p.r>r.left&&p.l<r.right){hit=p;break}
      }
      if(!hit)break;
      t=hit.b+2;
    }
    var dy=Math.round(t-r.top);
    if(dy>0){
      el.style.marginTop=(5+dy)+'px';
      el.insertAdjacentHTML('afterbegin',
        '<b class="fbl" style="height:'+(dy+5)+'px"></b>');
    }
    placed.push({l:r.left,r:r.right,t:t,b:t+r.height});
    low=Math.max(low,t+r.height);
  });
  var cv=sh.querySelector('.canvas');        // 下にはみ出した分だけ余白を足す
  sh.style.paddingBottom='';
  if(cv){
    var over=low-cv.getBoundingClientRect().bottom;
    if(over>0)sh.style.paddingBottom=Math.ceil(over+16)+'px';
  }
}
document.querySelectorAll('.tabs button').forEach(function(b){
  b.addEventListener('click',function(){
    document.querySelectorAll('.tabs button').forEach(function(x){x.classList.remove('on')});
    document.querySelectorAll('.sheet').forEach(function(x){x.classList.remove('on')});
    b.classList.add('on');
    var sh=document.getElementById('sh'+b.dataset.i);
    sh.classList.add('on'); mark(sh); place(sh);
  });
});
var cb=document.getElementById('markclip');
if(cb)cb.addEventListener('change',function(){
  document.body.classList.toggle('mark',cb.checked);
});
var gl=document.getElementById('gridlines');
if(gl)gl.addEventListener('change',function(){
  document.body.classList.toggle('nogl',!gl.checked);
});
var fb=document.getElementById('fbshow');
if(fb)fb.addEventListener('change',function(){
  document.body.classList.toggle('nofb',!fb.checked);
  if(fb.checked)place(document.querySelector('.sheet.on'));
});
var fd=document.getElementById('fbdup');
if(fd)fd.addEventListener('change',function(){
  document.body.classList.toggle('fbdup',fd.checked);
  place(document.querySelector('.sheet.on'));
});
var fm=document.getElementById('fxmark');
if(fm)fm.addEventListener('change',function(){
  document.body.classList.toggle('fxm',fm.checked);
});
// クリックで開閉。ドラッグ（範囲選択）と区別するため、押した位置から動いた場合と
// 文字が選択されている場合は無視する。
var px=0,py=0;
document.addEventListener('mousedown',function(e){px=e.clientX;py=e.clientY},true);
document.addEventListener('click',function(e){
  if(!e.target.closest)return;
  var el=e.target.closest('.fb')||e.target.closest('td>i.on');
  if(!el)return;
  if(Math.abs(e.clientX-px)>4||Math.abs(e.clientY-py)>4)return;
  var s=window.getSelection();
  if(s&&!s.isCollapsed)return;
  el.classList.toggle('open');
});
document.addEventListener('keydown',function(e){
  if(e.key==='Escape')document.querySelectorAll('td>i.open,.fb.open').forEach(function(el){
    el.classList.remove('open')});
});
// タブは画面下端に固定なので、本文の下にその高さぶんの余白を空ける
function fitTabs(){
  var t=document.querySelector('.tabs');
  if(t)document.body.style.paddingBottom=t.offsetHeight+'px';
}
window.addEventListener('resize',fitTabs);
fitTabs();
mark(document.querySelector('.sheet.on'));
place(document.querySelector('.sheet.on'));
"""
    toggle = ('<label class="tg" title="Excel と同じ位置で文字が切れているセルに印を付けます。'
              '印の有無にかかわらず、クリックすると全文を表示できます">'
              '<input type="checkbox" id="markclip">見切れに印</label>')
    if any(s.get("gridlines") for s in sheets):
        toggle = ('<label class="tg" title="Excel の目盛線（薄いグレーの線）の表示を切り替えます。'
                  '印刷には元から出ません">'
                  '<input type="checkbox" id="gridlines" checked>目盛線</label>') + toggle
    if any(s.get("balloons") for s in sheets):
        toggle += ('<label class="tg" title="数式セルの下に数式を吹き出しで表示します。'
                   '吹き出しにマウスを乗せると畳まれた分も含めて全文が出ます'
                   '（クリックで固定、Esc で全部閉じる）">'
                   '<input type="checkbox" id="fbshow" checked>数式の吹き出し</label>')
        if any(s.get("fbdup") for s in sheets):
            toggle += ('<label class="tg" title="上や左のセルと同じ式（コピーされた式）は、'
                       '既定では先頭のセルだけに吹き出しを出し「×何セル分」と添えています。'
                       'これをオンにすると、繰り返しの分も全部出します">'
                       '<input type="checkbox" id="fbdup">同じ式も出す</label>')
        toggle += ('<label class="tg" title="数式が入っているセルの右上に赤い三角を付けます">'
                   '<input type="checkbox" id="fxmark">数式セルに印</label>')
    tabs_html = f'<div class="tabs">{tabs if len(sheets) > 1 else ""}{toggle}</div>'
    return (f'<!doctype html>\n<html lang="ja"><head><meta charset="utf-8">'
            f'<meta name="viewport" content="width=device-width,initial-scale=1">'
            f"<title>{esc(title)}</title>\n<style>\n{''.join(css)}\n</style></head>\n"
            f'<body><div class="wrap">{bodies}</div>{tabs_html}'
            f"<script>{script}</script></body></html>\n")


def check_container(src: str):
    """xlsx(zip) でないファイルを分かりやすく弾く。"""
    with open(src, "rb") as f:
        head = f.read(8)
    if head[:4] == b"\xd0\xcf\x11\xe0":
        raise SystemExit(
            "このファイルは xlsx ではありません（旧 .xls 形式、またはパスワード保護／IRM 付き）。\n"
            "Excel で開いて『名前を付けて保存 → Excel ブック(.xlsx)』で保存し直してから再実行してください。")
    if head[:2] != b"PK":
        raise SystemExit("xlsx として読めない形式です。")


_FUNC_RE = re.compile(r"([A-Z][A-Z0-9_.]*)\s*\(")


def diagnose(src: str, opts) -> None:
    """数式まわりの状態だけを調べて表示する。

    セルの中身（値・文字列）は一切出さない。出すのは件数と、関数名・セル番地だけ。
    機密ファイルでも安全に実行して結果を共有できるようにしてある。
    """
    check_container(src)
    print(f"■ 診断: {src}\n")

    with zipfile.ZipFile(src) as z:
        names = set(z.namelist())
        try:
            app = ET.fromstring(z.read("docProps/app.xml"))
            gen = "".join(e.text or "" for e in app if e.tag.endswith("}Application"))
            ver = "".join(e.text or "" for e in app if e.tag.endswith("}AppVersion"))
            print(f"作成アプリ : {gen or '不明'} {ver}")
        except (KeyError, ET.ParseError):
            print("作成アプリ : 不明（docProps/app.xml なし＝Excel 以外の生成物の可能性）")
        try:
            wbx = ET.fromstring(z.read("xl/workbook.xml"))
            cp = wbx.find(f"{{{MAIN}}}calcPr")
            if cp is not None:
                print(f"計算設定   : calcId={cp.get('calcId')} "
                      f"calcMode={cp.get('calcMode') or 'auto'} "
                      f"fullCalcOnLoad={cp.get('fullCalcOnLoad') or '0'}")
                if cp.get("fullCalcOnLoad") in ("1", "true"):
                    print("             → Excel は開くたび全再計算する設定。"
                          "保存済みの計算結果は古い可能性が高い")
                if (cp.get("calcMode") or "auto") == "manual":
                    print("             → 手動計算。保存値が最新とは限らない")
            else:
                print("計算設定   : calcPr なし")
        except (KeyError, ET.ParseError):
            pass
        print(f"calcChain  : {'あり' if 'xl/calcChain.xml' in names else 'なし'}"
              f"（なし＝Excel が計算結果を書いていない可能性）")

    wb_v = load_workbook(src, data_only=True)
    wb_f = load_workbook(src, data_only=False, read_only=True)
    targets = wb_v.sheetnames if not opts.sheet else [opts.sheet]
    try:
        from openpyxl.worksheet.formula import ArrayFormula
    except ImportError:
        ArrayFormula = ()

    funcs: dict[str, int] = {}
    tot = {"f": 0, "none": 0, "err": 0, "num": 0, "str": 0, "array": 0}
    for name in targets:
        if name not in wb_v.sheetnames:
            continue
        wsv, wsf = wb_v[name], wb_f[name]
        per_col: dict[int, list[tuple[int, object]]] = {}
        n_f = n_none = n_err = n_arr = 0
        for row in wsf.iter_rows():
            for cell in row:
                fv = cell.value
                if ArrayFormula and isinstance(fv, ArrayFormula):
                    fv, n_arr = fv.text, n_arr + 1
                if not (isinstance(fv, str) and fv.startswith("=")):
                    continue
                n_f += 1
                for fn in _FUNC_RE.findall(tidy_formula(fv).upper()):
                    funcs[fn] = funcs.get(fn, 0) + 1
                v = wsv.cell(row=cell.row, column=cell.column).value
                if v is None:
                    n_none += 1
                elif isinstance(v, str) and v in ERROR_TEXTS:
                    n_err += 1
                    tot["err"] += 1
                else:
                    tot["num" if isinstance(v, (int, float)) else "str"] += 1
                per_col.setdefault(cell.column, []).append((cell.row, v))
        tot["f"] += n_f
        tot["none"] += n_none
        tot["array"] += n_arr
        if not n_f:
            continue
        print(f"\n── シート '{name}'")
        print(f"   数式セル {n_f} 個 / 計算結果なし {n_none} 個 / エラー値 {n_err} 個"
              + (f" / 配列数式 {n_arr} 個" if n_arr else ""))

        # 「同じ計算結果ばかり」の検出: 縦に連続する数式セルの保存値が同一な区間
        worst = None
        n_runs = 0
        for col, items in per_col.items():
            items.sort()
            run_start, run_len = 0, 1
            for i in range(1, len(items) + 1):
                same = (i < len(items) and items[i][0] == items[i - 1][0] + 1
                        and items[i][1] == items[i - 1][1] and items[i][1] is not None)
                if same:
                    run_len += 1
                    continue
                if run_len >= 3:
                    n_runs += 1
                    if worst is None or run_len > worst[0]:
                        worst = (run_len, col, items[run_start][0])
                run_start, run_len = i, 1
        if worst:
            a = f"{get_column_letter(worst[1])}{worst[2]}"
            b = f"{get_column_letter(worst[1])}{worst[2] + worst[0] - 1}"
            print(f"   ※ 縦に連続する数式セルで保存値が同一の区間が {n_runs} か所"
                  f"（最長 {worst[0]} セル: {a}:{b}）")
            print("     → Excel の画面で同じ範囲がどう見えるか比べてください。"
                  "Excel 側も同じ値なら変換は正しく、")
            print("       Excel 側が違う値なら、xlsx に保存された計算結果自体が古いということです。")
        else:
            print("   保存値が縦に不自然な形で揃っている箇所はありません")

    diagnose_formats(wb_v, targets)

    print(f"\n── 全体: 数式 {tot['f']} 個"
          f" / 計算結果なし {tot['none']} 個 / エラー値 {tot['err']} 個"
          f" / 数値 {tot['num']} 個 / 文字列 {tot['str']} 個")
    if funcs:
        top = sorted(funcs.items(), key=lambda kv: -kv[1])[:25]
        print("   使われている関数: " + ", ".join(f"{k}×{v}" for k, v in top))
    if tot["none"]:
        print("   ※ 計算結果が保存されていないセルがあります。"
              "Excel で開いて上書き保存すると解消します。", file=sys.stderr)
    wb_v.close()
    wb_f.close()


def diagnose_formats(wb, targets: list[str]) -> None:
    """日付が数値のまま残っていないかを表示形式ごとに数える。

    出すのは numFmtId と件数だけで、セルの中身も書式の文字列も出さない。
    """
    if is_date_format is None:
        return
    dated: dict[int, int] = {}          # numFmtId -> 日付として読めたセル数
    raw: dict[int, int] = {}            # numFmtId -> 日付書式なのに数値のままのセル数
    unknown: dict[int, int] = {}        # 表示形式そのものを解決できない numFmtId
    for name in targets:
        if name not in wb.sheetnames:
            continue
        for row in wb[name].iter_rows():
            for cell in row:
                v = cell.value
                if v is None or isinstance(v, (bool, str)):
                    continue
                nid = getattr(getattr(cell, "_style", None), "numFmtId", 0)
                if isinstance(v, (datetime, date, time, timedelta)):
                    dated[nid] = dated.get(nid, 0) + 1
                elif isinstance(v, (int, float)):
                    if nid < 164 and nid not in _opx_numbers.BUILTIN_FORMATS:
                        unknown[nid] = unknown.get(nid, 0) + 1
                    elif is_date_format(cell.number_format):
                        raw[nid] = raw.get(nid, 0) + 1

    def ids(d: dict[int, int]) -> str:
        return ", ".join(f"{k}({v}個)" for k, v in sorted(d.items()))

    print("\n── 表示形式")
    print(f"   日付・時刻として読めたセル: {sum(dated.values())} 個"
          + (f"  numFmtId {ids(dated)}" if dated else ""))
    if raw:
        print(f"   ※ 日付書式なのに数値のまま残っているセル: {sum(raw.values())} 個"
              f"  numFmtId {ids(raw)}")
        print("     → 45000 前後のシリアル値がそのまま表示されます。この numFmtId を教えてください。")
    if unknown:
        print(f"   ※ 表示形式を解決できない numFmtId: {ids(unknown)}")
        print("     → General 扱いになります。この numFmtId を教えてください。")


def convert(src: str, dst: str, opts) -> str:
    check_container(src)
    wb = load_workbook(src, data_only=True, rich_text=True)
    resolver = ColorResolver(getattr(wb, "loaded_theme", None))
    parser = None
    if not opts.no_drawings:
        try:
            parser = DrawingParser(src, resolver)
        except Exception as e:      # 図形解析の失敗で全体を落とさない
            print(f"警告: 図形の解析に失敗しました ({e})", file=sys.stderr)
            parser = None

    targets = wb.sheetnames if not opts.sheet else [opts.sheet]
    formulas: dict[str, dict] = {}
    if (opts.show_formula or opts.formula_tips or opts.formula_balloons
            or not opts.no_formula_check):
        try:
            formulas = load_formulas(src, targets)
        except Exception as e:
            print(f"警告: 数式の読み込みに失敗しました ({e})", file=sys.stderr)
    n_formula = sum(len(m) for m in formulas.values())
    missing_total = 0

    # 保護ビューのせいで #VALUE! が保存されているシート名の式を解決しておく
    fixed: dict[str, dict[tuple[int, int], str]] = {}
    if formulas:
        try:
            fixed = resolve_sheet_name_formulas(wb, formulas, os.path.basename(src))
        except Exception as e:      # 解決できなくても保存値をそのまま出せばよい
            print(f"警告: シート名の式を解決できませんでした ({e})", file=sys.stderr)
    n_fixed = sum(len(m) for m in fixed.values())
    n_cf = n_cf_cells = 0
    unsupported_cf: set[str] = set()
    uneval_cf: set[str] = set()

    sheets = []
    for name in targets:
        if name not in wb.sheetnames:
            print(f"警告: シート '{name}' が見つかりません", file=sys.stderr)
            continue
        ws = wb[name]
        if ws.sheet_state != "visible" and not opts.hidden_sheets:
            continue
        if opts.gridlines == "on":
            gl = True
        elif opts.gridlines == "off":
            gl = False
        else:
            # 属性が省略されている場合、OOXML の既定は「表示する」
            sv = getattr(ws.sheet_view, "showGridLines", None)
            gl = True if sv is None else bool(sv)
        max_row = min(ws.max_row or 1, opts.max_rows)
        max_col = min(ws.max_column or 1, opts.max_cols)
        grid_for_shapes = Grid(ws, max_col, max_row)
        shapes = []
        if parser is not None:
            try:
                shapes = parser.shapes_for(name, grid_for_shapes)
            except Exception as e:
                print(f"警告: '{name}' の図形を読み込めませんでした ({e})", file=sys.stderr)
        bounds = print_area_bounds(ws, opts) if opts.print_area else None
        breaks = {b.id + 1 for b in ws.row_breaks.brk} if ws.row_breaks else set()
        cond = None
        if not opts.no_cond_format:
            cb = bounds or (1, 1, max_row, max_col)
            try:
                cond = CondFormat(wb, ws, resolver, cb)
            except Exception as e:      # 条件付き書式で全体を落とさない
                print(f"警告: '{name}' の条件付き書式を評価できませんでした ({e})",
                      file=sys.stderr)
                cond = None
        res = render_sheet(ws, resolver, shapes, gl, name, opts, formulas.get(name),
                           bounds, breaks, title_row_range(ws), fixed.get(name), cond)
        missing_total += res["missing"]
        if cond is not None and cond.n_rules:
            n_cf += cond.n_rules
            n_cf_cells += len(cond.map)
            if cond.truncated:
                print(f"警告: '{name}' は条件付き書式の対象セルが多すぎるため一部を省略しました",
                      file=sys.stderr)
            unsupported_cf.update(cond.unsupported)
            uneval_cf.update(cond.uneval)
        page = page_settings(ws, res["width"]) if not opts.no_page_setup else None
        po = getattr(ws, "print_options", None)
        sheets.append({"name": name, "html": res["html"], "css": res["css"], "page": page,
                       "gridlines": gl, "balloons": res["balloons"], "fbdup": res["fbdup"],
                       "print_gridlines": bool(gl and po is not None and po.gridLines)})
        note = f", 計算結果なしの数式{res['missing']}個" if res["missing"] else ""
        area = "（印刷範囲で切り出し）" if bounds else ""
        print(f"  ・{name}: {res['cols']}列 × {res['rows']}行, 図形{len(shapes)}個{note}{area}")
    if parser is not None:
        parser.close()
    if not sheets:
        raise SystemExit("変換対象のシートがありません")

    if n_formula:
        print(f"  数式セル {n_formula} 個 → 保存済みの計算結果を表示"
              + (f"（うち {missing_total} 個は結果が未保存）" if missing_total else ""))
    n_fb = sum(s["balloons"] for s in sheets)
    if n_fb:
        n_dup = sum(s["fbdup"] for s in sheets)
        print(f"  数式の吹き出し {n_fb} 個"
              + (f"（うち {n_dup} 個は上や左と同じ式なので先頭以外は隠しています。"
                 "画面下部の「同じ式も出す」で全部出せます）" if n_dup else ""))
    if n_fixed:
        print(f"  シート名を出す数式 {n_fixed} 個 → 保存されていたエラーを解決して表示")
    if n_cf:
        print(f"  条件付き書式 {n_cf} ルール → {n_cf_cells} セルに反映")
    if unsupported_cf:
        print(f"  ※ 未対応の条件付き書式: {', '.join(sorted(unsupported_cf))}", file=sys.stderr)
    if uneval_cf:
        print(f"  ※ 判定できなかった数式ルール {len(uneval_cf)} 件（そのルールは不成立として扱いました）:",
              file=sys.stderr)
        for u in sorted(uneval_cf)[:5]:
            print(f"      {u}", file=sys.stderr)
    if missing_total and not opts.show_formula:
        print("  ※ 計算結果が保存されていない数式セルは空欄になります。"
              "Excel で開いて上書き保存するか、--show-formula で数式そのものを出せます。",
              file=sys.stderr)

    title = os.path.splitext(os.path.basename(src))[0]
    out = build_html(title, sheets)
    with open(dst, "w", encoding="utf-8", newline="\n") as f:
        f.write(out)
    return dst


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Excel 方眼紙ドキュメントを見た目そのままの HTML に変換します。")
    ap.add_argument("input", help="入力 .xlsx / .xlsm")
    ap.add_argument("-o", "--output", help="出力 .html（既定: 入力と同名）")
    ap.add_argument("--sheet", help="変換するシート名（既定: 全シート）")
    ap.add_argument("--gridlines", choices=["auto", "on", "off"], default="auto",
                    help="目盛線の描画（既定 auto = Excel の設定に従う）")
    ap.add_argument("--no-drawings", action="store_true", help="画像・図形を出力しない")
    ap.add_argument("--no-cond-format", action="store_true",
                    help="条件付き書式（色付け・データバー・アイコン）を反映しない")
    ap.add_argument("--diagnose", action="store_true",
                    help="変換せず、数式まわりの状態だけを調べて表示する")
    ap.add_argument("--hidden-sheets", action="store_true", help="非表示シートも出力する")
    ap.add_argument("--max-cols", type=int, default=1024, help="出力する最大列数")
    ap.add_argument("--max-rows", type=int, default=20000, help="出力する最大行数")
    fg = ap.add_mutually_exclusive_group()
    fg.add_argument("--formula-tips", action="store_true",
                    help="数式セルにマウスを乗せると数式を表示する")
    fg.add_argument("--formula-balloons", action="store_true",
                    help="数式セルの下に数式を吹き出しで常時表示する（数式チェック用）")
    ap.add_argument("--show-formula", action="store_true",
                    help="計算結果が保存されていない数式セルに、数式そのものを表示する")
    ap.add_argument("--no-formula-check", action="store_true",
                    help="数式の走査を省略して高速化する")
    ap.add_argument("--print-area", action="store_true",
                    help="印刷範囲が設定されているシートは、その範囲だけを出力する")
    ap.add_argument("--no-page-setup", action="store_true",
                    help="用紙サイズ・余白・改ページなどの印刷設定を反映しない")
    opts = ap.parse_args(argv)

    src = opts.input
    if not os.path.exists(src):
        raise SystemExit(f"入力が見つかりません: {src}")
    if opts.diagnose:
        diagnose(src, opts)
        return
    dst = opts.output or os.path.splitext(src)[0] + ".html"
    print(f"変換中: {src}")
    convert(src, dst, opts)
    size = os.path.getsize(dst)
    print(f"完了: {dst}  ({size/1024:.1f} KB)")


if __name__ == "__main__":
    main()
