#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_sample.py — 動作検証用の「Excel 方眼紙」サンプルを生成する。

日本の社内資料でよくある要素をひととおり詰め込む:
  細かい方眼 / 結合セルの見出し / 太罫線の枠 / 表示形式(日付・金額・%) /
  折り返し / 縦書き / 斜線 / セル内一部書式 / 画像 / オートシェイプ + 矢印
オートシェイプは openpyxl で作れないため、保存後に drawing XML へ直接注入する。
"""

import io
import os
import re
import shutil
import zipfile

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_houganshi.xlsx")

THIN = Side(style="thin", color="808080")
MED = Side(style="medium", color="000000")
DOT = Side(style="dotted", color="999999")


def box(ws, r1, c1, r2, c2, outer=MED, inner=None):
    """範囲に外枠(+内側罫線)を引く。"""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            b = cell.border
            left = outer if c == c1 else (inner or b.left)
            right = outer if c == c2 else (inner or b.right)
            top = outer if r == r1 else (inner or b.top)
            bottom = outer if r == r2 else (inner or b.bottom)
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def put(ws, r, c, value, *, span=None, font=None, align=None, fill=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=value)
    if span:
        ws.merge_cells(start_row=r, start_column=c,
                       end_row=r + span[0] - 1, end_column=c + span[1] - 1)
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        cell.number_format = fmt
    return cell


def make_logo(path):
    img = Image.new("RGBA", (180, 180), (0, 0, 0, 0))
    d = ImageDraw.Draw(img)
    d.ellipse((5, 5, 175, 175), fill=(16, 124, 65, 255))
    d.polygon([(55, 90), (85, 125), (135, 55)], fill=(255, 255, 255, 255))
    d.ellipse((5, 5, 175, 175), outline=(10, 80, 42, 255), width=6)
    img.save(path)


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "作業報告書"

    # --- 方眼の設定: 幅 2.13(=20px) / 高さ 15pt(=20px) ---
    NCOL, NROW = 46, 46
    ws.sheet_format.defaultRowHeight = 15
    for c in range(1, NCOL + 1):
        ws.column_dimensions[get_column_letter(c)].width = 2.13
    for r in range(1, NROW + 1):
        ws.row_dimensions[r].height = 15
    ws.sheet_view.showGridLines = True

    gothic = "Meiryo"
    f_title = Font(name=gothic, size=18, bold=True, color="1F3864")
    f_h = Font(name=gothic, size=9, bold=True, color="FFFFFF")
    f_lbl = Font(name=gothic, size=9, bold=True)
    f = Font(name=gothic, size=9)
    f_small = Font(name=gothic, size=8, color="666666")
    ctr = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # --- タイトル ---
    put(ws, 2, 2, "月 次 作 業 報 告 書", span=(3, 30), font=f_title, align=ctr)
    put(ws, 2, 33, "文書番号", span=(1, 3), font=f_small, align=right)
    put(ws, 2, 36, "QC-2026-0731", span=(1, 5), font=f_small, align=left)
    put(ws, 3, 33, "作成日", span=(1, 3), font=f_small, align=right)
    c = put(ws, 3, 36, __import__("datetime").date(2026, 7, 26), span=(1, 5),
            font=f_small, align=left, fmt='yyyy"年"m"月"d"日"(aaa)')

    # --- ヘッダ情報（左右2ブロック） ---
    rows = [("部　　署", "第一システム部 / 基盤グループ"),
            ("担 当 者", "山田 太郎"),
            ("対象期間", "2026/07/01 〜 2026/07/31")]
    r0 = 6
    for i, (k, v) in enumerate(rows):
        rr = r0 + i * 2
        put(ws, rr, 2, k, span=(2, 6), font=f_lbl, align=ctr, fill="DDEBF7")
        put(ws, rr, 8, v, span=(2, 14), font=f, align=left)
        box(ws, rr, 2, rr + 1, 7, outer=THIN)
        box(ws, rr, 8, rr + 1, 21, outer=THIN)

    put(ws, 6, 24, "承　認", span=(2, 5), font=f_lbl, align=ctr, fill="FCE4D6")
    for i, k in enumerate(["部長", "課長", "担当"]):
        put(ws, 6, 29 + i * 4, k, span=(1, 4), font=f_small, align=ctr, fill="F2F2F2")
        box(ws, 6, 29 + i * 4, 11, 32 + i * 4, outer=THIN)
    box(ws, 6, 24, 11, 28, outer=THIN)
    box(ws, 6, 24, 11, 40, outer=MED)

    # --- 明細表 ---
    top = 14
    put(ws, top - 1, 2, "■ 作業実績", span=(1, 10), font=Font(name=gothic, size=10, bold=True,
                                                              color="1F3864"), align=left)
    headers = [("No.", 3), ("作業項目", 13), ("工数(h)", 5), ("単価", 6),
               ("金額", 7), ("進捗", 4), ("完了予定", 6)]
    col = 2
    for name, w in headers:
        put(ws, top, col, name, span=(1, w), font=f_h, align=ctr, fill="4472C4")
        col += w
    data = [
        (1, "要件定義・ヒアリング", 24, 8500, 0.30, "2026/08/05"),
        (2, "基本設計書レビュー", 16.5, 9000, 1.00, "2026/07/18"),
        (3, "DB マイグレーション実装", 42, 7800, 0.75, "2026/08/12"),
        (4, "CI/CD パイプライン整備", 8, 9500, 1.00, "2026/07/22"),
        (5, "障害調査（緊急対応）", 5.5, 12000, 0.10, "2026/08/31"),
    ]
    import datetime as _dt
    for i, (no, item, hrs, unit, prog, due) in enumerate(data):
        r = top + 1 + i * 2
        put(ws, r, 2, no, span=(2, 3), font=f, align=ctr)
        put(ws, r, 5, item, span=(2, 13), font=f, align=left)
        put(ws, r, 18, hrs, span=(2, 5), font=f, align=right, fmt="#,##0.0")
        put(ws, r, 23, unit, span=(2, 6), font=f, align=right, fmt='"¥"#,##0')
        put(ws, r, 29, hrs * unit, span=(2, 7), font=f, align=right, fmt='"¥"#,##0;[赤]-"¥"#,##0')
        put(ws, r, 36, prog, span=(2, 4), font=f, align=ctr, fmt="0%")
        put(ws, r, 40, _dt.datetime.strptime(due, "%Y/%m/%d").date(), span=(2, 6),
            font=f, align=ctr, fmt="m/d")
        box(ws, r, 2, r + 1, 45, outer=THIN, inner=None)
        for cc, w in [(2, 3), (5, 13), (18, 5), (23, 6), (29, 7), (36, 4), (40, 6)]:
            box(ws, r, cc, r + 1, cc + w - 1, outer=THIN)

    # 合計行
    rt = top + 1 + len(data) * 2
    put(ws, rt, 2, "合　計", span=(2, 16), font=f_lbl, align=ctr, fill="F2F2F2")
    put(ws, rt, 18, sum(d[2] for d in data), span=(2, 5), font=f_lbl, align=right, fmt="#,##0.0")
    put(ws, rt, 23, "", span=(2, 6), font=f, align=right)
    put(ws, rt, 29, sum(d[2] * d[3] for d in data), span=(2, 7), font=Font(
        name=gothic, size=10, bold=True, color="C00000"), align=right, fmt='"¥"#,##0')
    put(ws, rt, 36, "", span=(2, 10), font=f)
    box(ws, rt, 2, rt + 1, 45, outer=THIN)
    for cc, w in [(2, 16), (18, 5), (23, 6), (29, 7), (36, 4), (40, 6)]:
        box(ws, rt, cc, rt + 1, cc + w - 1, outer=THIN)
    box(ws, top, 2, rt + 1, 45, outer=MED)

    # 表ヘッダの下線を太く
    for cc in range(2, 46):
        cell = ws.cell(row=top, column=cc)
        b = cell.border
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=MED)

    # --- 備考（折り返し + 一部書式） ---
    rb = rt + 3
    put(ws, rb, 2, "備\n考", span=(6, 2), font=f_lbl,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True), fill="F2F2F2")
    note = ("・障害調査は 7/29 に再発生したため、8 月へ持ち越し。根本原因は接続プールの枯渇と推定。\n"
            "・DB マイグレーションは検証環境まで完了。本番反映は変更諮問委員会の承認待ち。\n"
            "・次月は SLO 定義とダッシュボード整備に着手する予定。")
    put(ws, rb, 4, note, span=(6, 30), font=f, align=wrap)
    box(ws, rb, 2, rb + 5, 3, outer=THIN)
    box(ws, rb, 4, rb + 5, 33, outer=THIN)

    # 縦書きラベル + 斜線セル
    put(ws, rb, 35, "社内限", span=(6, 2), font=Font(name=gothic, size=9, bold=True, color="C00000"),
        align=Alignment(horizontal="center", vertical="center", textRotation=255))
    box(ws, rb, 35, rb + 5, 36, outer=THIN)
    diag = ws.cell(row=rb, column=38)
    ws.merge_cells(start_row=rb, start_column=38, end_row=rb + 5, end_column=40)
    diag.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN,
                         diagonal=Side(style="thin", color="808080"), diagonalDown=True)

    # --- はみ出し文字のテスト（方眼紙で最頻出のパターン） ---
    put(ws, rb + 7, 2, "※ 本報告書は社内利用に限ります。1セルに入力した文章が右へはみ出す例です。",
        font=f_small, align=left)
    put(ws, rb + 9, 2, "※ 右隣にセル内容があるため、Excel と同じくここで文字が途切れます……"
                       "この先は表示されません。", font=f_small, align=left)
    put(ws, rb + 9, 22, "←ここで停止", font=f_small, align=left)

    # --- 印刷設定（A4 横 / 印刷範囲 / 印刷タイトル行 / 手動改ページ） ---
    ws.print_area = "B2:AS40"
    ws.print_title_rows = "2:4"
    ws.page_setup.paperSize = 9          # A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.5
    ws.page_margins.top = ws.page_margins.bottom = 0.6
    ws.row_breaks.append(__import__("openpyxl").worksheet.pagebreak.Break(id=26))

    # --- 画像 ---
    logo = os.path.join(os.path.dirname(OUT), "_logo.png")
    make_logo(logo)
    img = XLImage(logo)
    img.width, img.height = 46, 46
    ws.add_image(img, "AR2")

    wb.save(OUT)
    os.remove(logo)
    return OUT


# ---------------------------------------------------------------------------
# オートシェイプの注入（openpyxl では図形を作れないため）
# ---------------------------------------------------------------------------

NSDECL = (' xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
          ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"')

SHAPES_XML = """
<xdr:twoCellAnchor>
  <xdr:from><xdr:col>26</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>38</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
  <xdr:to><xdr:col>42</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>43</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>
  <xdr:sp macro="" textlink="">
    <xdr:nvSpPr><xdr:cNvPr id="101" name="角丸四角形 1"/><xdr:cNvSpPr/></xdr:nvSpPr>
    <xdr:spPr>
      <a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>
      <a:prstGeom prst="roundRect"><a:avLst/></a:prstGeom>
      <a:solidFill><a:srgbClr val="FFF2CC"/></a:solidFill>
      <a:ln w="19050"><a:solidFill><a:srgbClr val="ED7D31"/></a:solidFill></a:ln>
    </xdr:spPr>
    <xdr:txBody>
      <a:bodyPr rtlCol="0" anchor="ctr"/><a:lstStyle/>
      <a:p><a:pPr algn="ctr"/><a:r>
        <a:rPr lang="ja-JP" sz="900" b="1"><a:solidFill><a:srgbClr val="C00000"/></a:solidFill>
        <a:ea typeface="Meiryo"/></a:rPr><a:t>要フォロー</a:t></a:r></a:p>
      <a:p><a:pPr algn="ctr"/><a:r>
        <a:rPr lang="ja-JP" sz="800"><a:ea typeface="Meiryo"/></a:rPr>
        <a:t>8/5 の会議で報告</a:t></a:r></a:p>
    </xdr:txBody>
  </xdr:sp>
  <xdr:clientData/>
</xdr:twoCellAnchor>
<xdr:twoCellAnchor>
  <xdr:from><xdr:col>22</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>40</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
  <xdr:to><xdr:col>26</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>40</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>
  <xdr:cxnSp macro="">
    <xdr:nvCxnSpPr><xdr:cNvPr id="102" name="直線矢印コネクタ 2"/><xdr:cNvCxnSpPr/></xdr:nvCxnSpPr>
    <xdr:spPr>
      <a:xfrm flipH="1"><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>
      <a:prstGeom prst="straightConnector1"><a:avLst/></a:prstGeom>
      <a:ln w="28575"><a:solidFill><a:srgbClr val="C00000"/></a:solidFill>
        <a:headEnd type="triangle"/></a:ln>
    </xdr:spPr>
  </xdr:cxnSp>
  <xdr:clientData/>
</xdr:twoCellAnchor>
<xdr:twoCellAnchor>
  <xdr:from><xdr:col>3</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>38</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
  <xdr:to><xdr:col>11</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>43</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>
  <xdr:sp macro="" textlink="">
    <xdr:nvSpPr><xdr:cNvPr id="103" name="楕円 3"/><xdr:cNvSpPr/></xdr:nvSpPr>
    <xdr:spPr>
      <a:xfrm rot="-1200000"><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>
      <a:prstGeom prst="ellipse"><a:avLst/></a:prstGeom>
      <a:noFill/>
      <a:ln w="28575"><a:solidFill><a:srgbClr val="C00000"/></a:solidFill></a:ln>
    </xdr:spPr>
    <xdr:txBody><a:bodyPr rtlCol="0" anchor="ctr"/><a:lstStyle/>
      <a:p><a:pPr algn="ctr"/><a:r>
        <a:rPr lang="ja-JP" sz="1100" b="1"><a:solidFill><a:srgbClr val="C00000"/></a:solidFill>
        <a:ea typeface="Meiryo"/></a:rPr><a:t>済</a:t></a:r></a:p>
    </xdr:txBody>
  </xdr:sp>
  <xdr:clientData/>
</xdr:twoCellAnchor>
"""


def inject_shapes(path):
    """保存済み xlsx の drawing1.xml に図形を追記する。"""
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/drawings/drawing1.xml":
                xml = data.decode("utf-8")
                # 挿入する要素自身に名前空間を宣言しておく（openpyxl は既定名前空間で書き出すため）
                add = SHAPES_XML.replace("<xdr:twoCellAnchor>", "<xdr:twoCellAnchor" + NSDECL + ">")
                for close in ("</xdr:wsDr>", "</wsDr>"):
                    if close in xml:
                        xml = xml.replace(close, add + close)
                        break
                data = xml.encode("utf-8")
            zout.writestr(item, data)
    os.replace(tmp, path)


if __name__ == "__main__":
    p = build()
    inject_shapes(p)
    print("生成:", p, f"({os.path.getsize(p)/1024:.1f} KB)")
